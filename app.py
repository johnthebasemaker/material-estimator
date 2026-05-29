"""
Smart Material Estimator · app.py (v3)
Run: streamlit run app.py
"""
import io, os, sys, sqlite3
from datetime import date, datetime
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import streamlit as st

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from validate_data     import clean_inventory, clean_recipe, clean_equipment
from allocation_engine import build_demand_matrix

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="Smart Material Estimator",
                   page_icon="🏗️", layout="wide",
                   initial_sidebar_state="expanded")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PATH_A   = os.path.join(BASE_DIR, "Materials_DetailsAvailable_Qty.xlsx")
PATH_B   = os.path.join(BASE_DIR, "For_1_SQM.xlsx")
PATH_C   = os.path.join(BASE_DIR, "Equipment.xlsx")
SHEET_A, SHEET_B, SHEET_C = "Materials", "LINING SYSTEM MATERIAL CONSM", "Data Input"
LOCATION_ORDER = ["Brown Field", "TRAIN J", "TRAIN K"]
DB_PATH = os.path.join(BASE_DIR, "sme_database.db")

# ─────────────────────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500;600;700&display=swap');
:root{
  /* Map to Streamlit's native theme to seamlessly adapt to Light/Dark mode */
  --bg0: var(--background-color);
  --bg1: var(--secondary-background-color);
  --bg2: var(--secondary-background-color);
  --bg3: var(--background-color);
  --bg4: var(--secondary-background-color);
  --border: rgba(128, 128, 128, 0.2);
  --border2: rgba(128, 128, 128, 0.3);
  
  /* Preserving accent colors */
  --amber:#F59E0B;--amber2:#FCD34D;--amber-bg:rgba(245,158,11,.15);
  --green:#10B981;--green-bg:rgba(16,185,129,.15);
  --red:#EF4444;--red-bg:rgba(239,68,68,.15);
  --orange:#F97316;--orange-bg:rgba(249,115,22,.15);
  --yellow:#EAB308;--yellow-bg:rgba(234,179,8,.15);
  --blue:#3B82F6;--blue-bg:rgba(59,130,246,.15);
  
  /* Adaptive Text Colors using CSS color-mix */
  --t0: var(--text-color);
  --t1: var(--text-color);
  --t2: color-mix(in srgb, var(--text-color) 80%, transparent);
  --t3: color-mix(in srgb, var(--text-color) 60%, transparent);
  --t4: color-mix(in srgb, var(--text-color) 45%, transparent);
  --t5: color-mix(in srgb, var(--text-color) 30%, transparent);
}
html,body,[class*="css"]{font-family:'Sora',sans-serif!important;background:var(--bg0)!important;color:var(--t1);}
.main .block-container{padding-top:0!important;padding-bottom:3rem;max-width:1480px;}

/* sidebar */
[data-testid="stSidebar"]{background:var(--bg1)!important;border-right:1px solid var(--border)!important;}
[data-testid="stSidebar"] *{font-family:'Sora',sans-serif!important;}

/* ── TRUE STICKY HEADER & TABS ── */
header[data-testid="stHeader"]{display:none!important;}
[data-testid="stAppViewContainer"]{padding-top:0!important;}

/* 1. Freeze the main title block */
.sticky-header-wrap{
    position:sticky;top:0;z-index:9999;
    background:var(--bg0);
    padding:.8rem 1.5rem .5rem;
    margin-bottom:0;
}

/* 2. Freeze the Tabs right below the header */
[data-testid="stTabs"] > div:first-of-type {
    position: sticky;
    top: 50px; /* Sits right under sticky-header-wrap */
    z-index: 9998;
    background: var(--bg0);
    padding: 0.5rem 0;
    border-bottom: 2px solid var(--border);
}

/* tabs styling */
[data-testid="stTabs"] [data-baseweb="tab-list"]{background:transparent;border-bottom:none;gap:0;padding:0;}
[data-testid="stTabs"] [data-baseweb="tab"]{font-family:'JetBrains Mono',monospace!important;font-size:.7rem;font-weight:700;letter-spacing:.1em;text-transform:uppercase;color:var(--t4)!important;padding:.7rem 1.3rem;border-bottom:3px solid transparent;transition:all .2s;}
[data-testid="stTabs"] [aria-selected="true"]{color:var(--amber)!important;border-bottom:3px solid var(--amber)!important;}

/* buttons */
.stButton>button{font-family:'JetBrains Mono',monospace!important;font-size:.68rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;background:var(--amber)!important;color:#000!important;border:none!important;border-radius:4px!important;padding:.48rem 1.1rem!important;transition:all .15s!important;}
.stButton>button:hover{background:#FBBF24!important;transform:translateY(-1px);box-shadow:0 4px 14px rgba(245,158,11,.3)!important;}

/* metrics */
[data-testid="stMetric"]{background:var(--bg2);border:1px solid var(--border);border-radius:8px;padding:.85rem 1rem!important;transition:all .2s;cursor:default;}
[data-testid="stMetric"]:hover{background:var(--bg3)!important;border-color:var(--amber)!important;box-shadow:0 0 0 1px var(--amber);transform:translateY(-1px);}
[data-testid="stMetricLabel"]{font-family:'JetBrains Mono',monospace!important;font-size:.6rem!important;letter-spacing:.12em;text-transform:uppercase;color:var(--t3)!important;}
[data-testid="stMetricValue"]{font-family:'JetBrains Mono',monospace!important;font-size:1.8rem!important;color:var(--t0)!important;}
[data-testid="stMetricDelta"]{font-size:.72rem!important;}

/* info boxes */
[data-testid="stInfo"]   {background:var(--blue-bg)!important;border-left:3px solid var(--blue)!important; color:var(--t1)!important;}
[data-testid="stSuccess"]{background:var(--green-bg)!important;border-left:3px solid var(--green)!important; color:var(--t1)!important;}
[data-testid="stWarning"]{background:var(--amber-bg)!important;border-left:3px solid var(--amber)!important; color:var(--t1)!important;}
[data-testid="stError"]  {background:var(--red-bg)!important;border-left:3px solid var(--red)!important; color:var(--t1)!important;}

/* expander */
[data-testid="stExpander"]{background:var(--bg2)!important;border:1px solid var(--border)!important;border-radius:6px!important;}
[data-testid="stExpander"] summary{color:var(--t1)!important;}

/* select */
[data-baseweb="select"]>div,[data-baseweb="input"]>div{background:var(--bg2)!important;border-color:var(--border2)!important;color:var(--t0)!important;}

/* dataframe */
[data-testid="stDataFrame"]{border:1px solid var(--border)!important;border-radius:6px;overflow:hidden;}
hr{border-color:var(--border)!important;margin:.8rem 0!important;}

/* ── Custom components ── */
.sec-hdr{font-family:'JetBrains Mono',monospace;font-size:.6rem;font-weight:700;letter-spacing:.18em;text-transform:uppercase;color:var(--t5);border-bottom:1px solid var(--border);padding-bottom:.3rem;margin-bottom:.8rem;}
.card{background:var(--bg2);border:1px solid var(--border);border-radius:8px;padding:1rem 1.2rem;margin-bottom:.6rem;}
.card-amber{border-left:4px solid var(--amber);}
.card-green{border-left:4px solid var(--green);}
.card-blue {border-left:4px solid var(--blue);}
.loc-badge{display:inline-block;font-family:'JetBrains Mono',monospace;font-size:.62rem;font-weight:700;letter-spacing:.07em;text-transform:uppercase;padding:.18rem .6rem;border-radius:3px;}
.loc-bf{background:rgba(59,130,246,.12);color:var(--blue);}
.loc-tj{background:rgba(245,158,11,.12);color:var(--amber);}
.loc-tk{background:rgba(16,185,129,.12);color:var(--green);}
.pill{display:inline-block;font-family:'JetBrains Mono',monospace;font-size:.68rem;font-weight:600;padding:.15rem .5rem;border-radius:20px;}
.pill-g{background:var(--green-bg);color:var(--green);}
.pill-y{background:var(--yellow-bg);color:var(--yellow);}
.pill-o{background:var(--orange-bg);color:var(--orange);}
.pill-r{background:var(--red-bg);color:var(--red);}
.tag-chip{display:inline-block;font-family:'JetBrains Mono',monospace;font-size:.7rem;background:var(--bg3);color:var(--amber);border:1px solid var(--border2);border-radius:4px;padding:.15rem .5rem;margin:.12rem;}
.syscode-block{background:var(--bg3);border:1px solid var(--border2);border-radius:6px;padding:.7rem .9rem;margin:.35rem 0;}
.syscode-hdr{display:flex;align-items:center;gap:.7rem;margin-bottom:.5rem;}
.code-badge{font-family:'JetBrains Mono',monospace;font-size:.68rem;font-weight:700;background:var(--bg4);color:var(--amber);border:1px solid var(--amber-bg);border-radius:4px;padding:.2rem .55rem;}
.session-equip{background:var(--bg2);border:1px solid var(--border);border-radius:8px;padding:.9rem 1rem;margin-bottom:.5rem;}
.drag-handle{font-size:1rem;color:var(--t5);cursor:grab;user-select:none;padding:.2rem .4rem;}
.grand-box{background:linear-gradient(135deg, var(--amber-bg) 0%, var(--bg2) 70%);border:1px solid var(--border2);border-left:4px solid var(--amber);border-radius:8px;padding:1.1rem 1.4rem;}
.status-dot-g::before{content:"●";color:var(--green);margin-right:.4rem;}
.status-dot-o::before{content:"●";color:var(--orange);margin-right:.4rem;}
.status-dot-y::before{content:"●";color:var(--yellow);margin-right:.4rem;}
.status-dot-r::before{content:"●";color:var(--red);margin-right:.4rem;}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# ADMIN LOGIN GATE  —  change credentials here
# ─────────────────────────────────────────────────────────────────────────────
_ADMIN_USER = "admin"
_ADMIN_PASS = "admin2026"

def _show_login():
    st.markdown("""
    <div style="display:flex;justify-content:center;align-items:center;
                min-height:60vh;flex-direction:column;gap:1rem;">
      <div style="font-family:'JetBrains Mono',monospace;font-size:1.6rem;
                  font-weight:700;color:var(--t0);letter-spacing:-.01em;
                  margin-bottom:.5rem;">🏗 Smart Material Estimator</div>
      <div style="font-size:.8rem;color:var(--t3);margin-bottom:1.5rem;">
        Please log in to continue</div>
    </div>""", unsafe_allow_html=True)
    col_l, col_m, col_r = st.columns([1,1,1])
    with col_m:
        user = st.text_input("Username", key="_login_user", placeholder="Enter username")
        pwd  = st.text_input("Password", type="password", key="_login_pass",
                             placeholder="Enter password")
        if st.button("🔐  Login", use_container_width=True, key="_login_btn"):
            if user == _ADMIN_USER and pwd == _ADMIN_PASS:
                st.session_state["_authenticated"] = True
                st.rerun()
            else:
                st.error("❌ Invalid credentials. Please try again.")

if "_authenticated" not in st.session_state:
    st.session_state["_authenticated"] = False
if not st.session_state["_authenticated"]:
    _show_login()
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def db_available():
    return os.path.exists(DB_PATH)


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADER  (SQLite when available, fallback to Excel)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner="Loading project data…")
def load_all():
    if db_available():
        conn = get_db()
        inv = pd.read_sql(
            "SELECT material_code AS Material_Code, material_name AS Material_Name, "
            "nature AS Nature, uom AS UOM, "
            "available_qty AS Available_Qty, ordered_qty AS Ordered_Qty "
            "FROM inventory", conn)

        recipe = pd.read_sql(
            "SELECT lining_system_code AS Lining_System_Code, "
            "lining_system_short_name AS Lining_System_Short_Name, "
            "lining_type AS Lining_Type, material_code AS Material_Code, "
            "material_description AS Material_Description, "
            "material_name AS Material_Name, for_1_sqm AS For_1_SQM, uom AS UOM "
            "FROM recipe", conn)

        equip_raw = pd.read_sql(
            "SELECT location AS Location, type AS Type, "
            "lining_system_code AS Lining_System_Code, "
            "lining_system_short_name AS Lining_System_Short_Name, "
            "lining_type AS Lining_Type, equipment_tag AS \"Equipment_Tag_No.\", "
            "name AS Name, description AS Description, "
            "material_spec AS \"Material Spec.\", design AS Design, "
            "surface_area_sqm AS Surface_Area_SQM, lining_systems AS \"Lining_System+\" "
            "FROM equipment", conn)

        sqm_prog = pd.read_sql(
            "SELECT equipment_tag AS \"Equipment_Tag_No.\", "
            "lining_system_code AS Lining_System_Code, "
            "original_sqm, done_sqm, "
            "(original_sqm - done_sqm) AS remaining_sqm "
            "FROM sqm_progress", conn)
        conn.close()

        equip_sc = (equip_raw
            .groupby(["Equipment_Tag_No.", "Lining_System_Code",
                      "Lining_System_Short_Name"], as_index=False)
            ["Surface_Area_SQM"].sum()
            .rename(columns={"Surface_Area_SQM": "Total_SQM_Original"}))
        equip_sc = equip_sc.merge(
            sqm_prog[["Equipment_Tag_No.", "Lining_System_Code",
                      "remaining_sqm", "done_sqm"]],
            on=["Equipment_Tag_No.", "Lining_System_Code"], how="left")
        equip_sc["remaining_sqm"] = equip_sc["remaining_sqm"].fillna(equip_sc["Total_SQM_Original"])
        equip_sc["done_sqm"]      = equip_sc["done_sqm"].fillna(0)
        equip_sc["Total_SQM"]     = equip_sc["remaining_sqm"]

    else:
        # Fallback: read from Excel using existing clean helpers
        from validate_data import clean_inventory, clean_recipe, clean_equipment
        df_a_raw = pd.read_excel(PATH_A, sheet_name=SHEET_A)
        df_b_raw = pd.read_excel(PATH_B, sheet_name=SHEET_B)
        df_c_raw = pd.read_excel(PATH_C, sheet_name=SHEET_C)
        inv    = clean_inventory(df_a_raw)
        recipe = clean_recipe(df_b_raw)
        equip_raw = clean_equipment(df_c_raw)

        inv_full = df_a_raw.copy()
        inv_full.columns = inv_full.columns.str.strip()
        inv_full["Material_Code"] = inv_full["Material_Code"].astype(str).str.strip()
        ordered_col = next((c for c in inv_full.columns
                            if c.strip() in ("Ordered_Qty","Balance To Be Received")), None)
        if ordered_col:
            inv_full[ordered_col] = pd.to_numeric(inv_full[ordered_col],errors="coerce").fillna(0)
            inv_ordered = inv_full.groupby("Material_Code",as_index=False).agg(
                Ordered_Qty=(ordered_col,"sum"))
            inv = inv.merge(inv_ordered,on="Material_Code",how="left")
            inv["Ordered_Qty"] = inv["Ordered_Qty"].fillna(0)
        else:
            inv["Ordered_Qty"] = 0.0

        raw = df_c_raw.copy()
        raw.columns = raw.columns.str.strip()
        raw = raw.dropna(subset=["Equipment_Tag_No.","Lining_System_Code"])
        raw["Equipment_Tag_No."] = raw["Equipment_Tag_No."].astype(str).str.strip()
        raw["Location"]           = raw["Location"].astype(str).str.strip()
        raw["Type"]               = raw["Type"].astype(str).str.strip()
        raw["Surface_Area_SQM"]   = pd.to_numeric(raw["Surface_Area_SQM"],errors="coerce")
        raw["Lining_System_Code"] = raw["Lining_System_Code"].astype(float).astype(int).astype(str)
        for col in ["Name","Description","Lining_System_Short_Name"]:
            if col in raw.columns:
                raw[col] = raw[col].astype(str).str.strip()
        equip_raw = raw

        equip_sc = (equip_raw
            .groupby(["Equipment_Tag_No.","Lining_System_Code","Lining_System_Short_Name"],
                     as_index=False)["Surface_Area_SQM"].sum()
            .rename(columns={"Surface_Area_SQM":"Total_SQM_Original"}))
        equip_sc["done_sqm"]      = 0.0
        equip_sc["remaining_sqm"] = equip_sc["Total_SQM_Original"]
        equip_sc["Total_SQM"]     = equip_sc["Total_SQM_Original"]

    # ── Demand matrix (uses remaining SQM) ───────────────────────────────
    dm = equip_sc.merge(recipe, on="Lining_System_Code", suffixes=("_e","_r"))
    dm["Demand_Qty"] = dm["For_1_SQM"] * dm["Total_SQM"]
    if "Lining_System_Short_Name_e" in dm.columns:
        dm = dm.rename(columns={"Lining_System_Short_Name_e":"Lining_System_Short_Name"})
        dm.drop(columns=["Lining_System_Short_Name_r"],inplace=True,errors="ignore")
    dm = dm[["Equipment_Tag_No.","Lining_System_Code","Lining_System_Short_Name",
             "Total_SQM","Material_Code","Material_Name","UOM","Demand_Qty"]]

    # ── Equipment master ─────────────────────────────────────────────────
    eq_master = equip_raw.groupby("Equipment_Tag_No.", as_index=False).agg(
        Name          =("Name",          "first"),
        Description   =("Description",   "first"),
        Location      =("Location",      "first"),
        Type          =("Type",          "first"),
        Lining_Systems=("Lining_System+","first"),
        Lining_Type   =("Lining_Type",   "first"),
        Material_Spec =("Material Spec.","first"),
        Design        =("Design",        "first"),
        Total_SQM     =("Surface_Area_SQM","sum"),
    )
    eq_master["Location"] = eq_master["Location"].str.strip()
    eq_master["Type"]     = eq_master["Type"].str.strip()

    # ── SQM reference (includes done_sqm for progress tracking) ─────────
    sqm_ref = equip_sc[["Equipment_Tag_No.","Lining_System_Code",
                         "Total_SQM","Total_SQM_Original","done_sqm"]].drop_duplicates()

    return inv, recipe, equip_sc, dm, eq_master, sqm_ref


inv, recipe, equip_sc, dm, eq_master, sqm_ref = load_all()
ALL_TAGS      = sorted(eq_master["Equipment_Tag_No."].tolist())
INV_POOL_INIT    = inv.set_index("Material_Code")["Available_Qty"].to_dict()
INV_ORDERED_INIT = inv.set_index("Material_Code")["Ordered_Qty"].to_dict()

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
if "session_tags" not in st.session_state:
    st.session_state.session_tags = []

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def loc_badge(loc):
    cls = {"Brown Field":"loc-bf","TRAIN J":"loc-tj","TRAIN K":"loc-tk"}.get(loc,"loc-bf")
    return f'<span class="loc-badge {cls}">{loc}</span>'

def status_dot(pct):
    if pct >= 100: return "status-dot-g"
    if pct >= 90:  return "status-dot-o"
    if pct >= 80:  return "status-dot-y"
    return "status-dot-r"

def fulfil_pill(pct):
    cls = "pill-g" if pct>=100 else "pill-o" if pct>=90 else "pill-y" if pct>=80 else "pill-r"
    return f'<span class="pill {cls}">{pct:.1f}%</span>'

@st.cache_data(show_spinner=False)
def _cached_cascade_allocate(tag_order_tuple: tuple) -> pd.DataFrame:
    """Cached worker for cascade_allocate"""
    pool = dict(INV_POOL_INIT)  # mutable copy
    rows = []
    for tag in tag_order_tuple:
        tag_dm = dm[dm["Equipment_Tag_No."] == tag].copy()
        # Process system codes in numeric order for consistency
        for code in sorted(tag_dm["Lining_System_Code"].unique(), key=lambda x: int(x)):
            code_rows = tag_dm[tag_dm["Lining_System_Code"] == code]
            for _, r in code_rows.iterrows():
                mat     = r["Material_Code"]
                demand  = r["Demand_Qty"]
                before  = pool.get(mat, 0.0)
                alloc   = min(demand, before)
                short   = demand - alloc
                after   = max(0.0, before - alloc)
                pool[mat] = after
                rows.append({
                    "Equipment_Tag_No.":       tag,
                    "Lining_System_Code":      code,
                    "Lining_System_Short_Name": r["Lining_System_Short_Name"],
                    "Total_SQM":               r["Total_SQM"],
                    "Material_Code":           mat,
                    "Material_Name":           r["Material_Name"],
                    "UOM":                     r["UOM"],
                    "Demand_Qty":              round(demand, 4),
                    "Allocated_Qty":           round(alloc, 4),
                    "Shortfall_Qty":           round(short, 4),
                    "Pool_Before":             round(before, 4),
                    "Pool_After":              round(after, 4),
                })
    result = pd.DataFrame(rows)
    if not result.empty:
        result["Fulfillment_Pct"] = (
            result["Allocated_Qty"] / result["Demand_Qty"].replace(0, np.nan) * 100
        ).fillna(100).clip(0, 100).round(2)
    return result

def cascade_allocate(tag_order: list[str]) -> pd.DataFrame:
    """
    Cascade inventory pool through equipment in order.
    Pool is GLOBAL per material (not per system code).
    """
    # Convert list to tuple so Streamlit's cache engine can hash it safely
    return _cached_cascade_allocate(tuple(tag_order))

def tag_fulfillment(alloc_df: pd.DataFrame, tag: str) -> float:
    t = alloc_df[alloc_df["Equipment_Tag_No."] == tag]
    if t.empty: return 0.0
    d = t["Demand_Qty"].sum()
    a = t["Allocated_Qty"].sum()
    return min(100.0, a / d * 100) if d > 0 else 100.0

def syscode_fulfillment(alloc_df: pd.DataFrame, tag: str, code: str) -> float:
    t = alloc_df[(alloc_df["Equipment_Tag_No."]==tag)&(alloc_df["Lining_System_Code"]==code)]
    if t.empty: return 0.0
    d = t["Demand_Qty"].sum()
    a = t["Allocated_Qty"].sum()
    return min(100.0, a / d * 100) if d > 0 else 100.0

def sqm_can_do(alloc_df: pd.DataFrame, tag: str, code: str) -> tuple[float, float, float]:
    """
    Returns (total_sqm, sqm_can_do, sqm_shortfall) for a (tag, system_code) pair.
    Method: weighted avg fulfillment % × Total_SQM
    """
    rows = alloc_df[(alloc_df["Equipment_Tag_No."]==tag) &
                    (alloc_df["Lining_System_Code"]==code)]
    if rows.empty:
        return 0.0, 0.0, 0.0
    total_sqm = sqm_ref[
        (sqm_ref["Equipment_Tag_No."]==tag) &
        (sqm_ref["Lining_System_Code"]==code)
    ]["Total_SQM"].sum()
    d = rows["Demand_Qty"].sum()
    a = rows["Allocated_Qty"].sum()
    pct = min(1.0, a / d) if d > 0 else 1.0
    can   = round(total_sqm * pct, 2)
    short = round(total_sqm - can, 2)
    return round(total_sqm, 2), can, short


def excel_bytes(df: pd.DataFrame, report_title: str = "",
                add_grand_total: bool = False) -> bytes:
    """
    Export df to Excel with:
    - Row 1: report_title (merged, bold, large font) if provided
    - Row 2: column headers (bold, background fill)
    - Rows 3+: data
    - Last row: GRAND TOTAL (bold, summing numeric columns) if add_grand_total=True
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    out_df = df.copy()

    # ── Grand total row ───────────────────────────────────────────────────
    if add_grand_total and len(out_df) > 0:
        num_cols = out_df.select_dtypes(include="number").columns.tolist()
        total_row = {c: "" for c in out_df.columns}
        total_row[out_df.columns[0]] = "GRAND TOTAL"
        for c in num_cols:
            try:
                total_row[c] = out_df[c].sum()
            except Exception:
                pass
        out_df = pd.concat([out_df, pd.DataFrame([total_row])], ignore_index=True)

    # ── Write to buffer ───────────────────────────────────────────────────
    if report_title:
        # Write title row first, then data
        title_df = pd.DataFrame([[report_title]])
        title_df.to_excel(buf, index=False, header=False, engine="openpyxl", startrow=0)
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            title_df.to_excel(w, index=False, header=False, startrow=0)
            out_df.to_excel(w, index=False, header=True, startrow=1)

        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active
        n_cols = len(out_df.columns)

        # Style title row (row 1)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=n_cols)
        title_cell = ws.cell(1, 1)
        title_cell.font      = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill      = PatternFill("solid", fgColor="1A2A3A")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Style header row (row 2)
        for col in range(1, n_cols + 1):
            cell = ws.cell(2, col)
            cell.font      = Font(bold=True, size=10, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="2D4A6A")
            cell.alignment = Alignment(horizontal="center")

        # Style grand total row (last row)
        if add_grand_total:
            last_row = ws.max_row
            for col in range(1, n_cols + 1):
                cell = ws.cell(last_row, col)
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill("solid", fgColor="F0C040")

        # Auto-width columns
        for col in range(1, n_cols + 1):
            max_len = max(
                (len(str(ws.cell(r, col).value or ""))
                 for r in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 40)

        out_buf = io.BytesIO()
        wb.save(out_buf)
        return out_buf.getvalue()
    else:
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            out_df.to_excel(w, index=False)
        return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# SUGGESTION ENGINE
# ─────────────────────────────────────────────────────────────────────────────
def _run_suggestion_engine(tag_list: list[str]) -> dict:
    """
    For each tag, tries moving it to every earlier position.
    Returns the best single-move gain per equipment and per system code.
    """
    if len(tag_list) < 2:
        return {"eq_suggestions": [], "sc_suggestions": [], "baseline_pcts": {}}

    baseline = cascade_allocate(tag_list)

    def _eq_pct(alloc, tag):
        t = alloc[alloc["Equipment_Tag_No."] == tag]
        d = t["Demand_Qty"].sum(); a = t["Allocated_Qty"].sum()
        return round(min(100.0, a/d*100), 2) if d > 0 else 100.0

    def _sc_pct(alloc, tag, code):
        t = alloc[(alloc["Equipment_Tag_No."]==tag)&(alloc["Lining_System_Code"]==code)]
        d = t["Demand_Qty"].sum(); a = t["Allocated_Qty"].sum()
        return round(min(100.0, a/d*100), 2) if d > 0 else 100.0

    baseline_pcts = {t: _eq_pct(baseline, t) for t in tag_list}
    eq_suggestions, sc_suggestions = [], []

    for i, target in enumerate(tag_list):
        if i == 0: continue
        base_pct = baseline_pcts[target]
        best_gain = 0; best_pos = i; best_new_pct = base_pct
        for new_pos in range(0, i):
            new_order = [t for t in tag_list if t != target]
            new_order.insert(new_pos, target)
            alloc   = cascade_allocate(new_order)
            new_pct = _eq_pct(alloc, target)
            gain    = new_pct - base_pct
            if gain > best_gain:
                best_gain = gain; best_pos = new_pos; best_new_pct = new_pct
        if best_gain > 0.4:
            eq_suggestions.append({
                "tag": target, "current_pos": i+1, "suggest_pos": best_pos+1,
                "current_pct": base_pct, "new_pct": best_new_pct,
                "gain": round(best_gain, 1),
            })

    for i, tag in enumerate(tag_list):
        if i == 0: continue
        tag_dm = dm[dm["Equipment_Tag_No."] == tag]
        for code in sorted(tag_dm["Lining_System_Code"].unique(), key=lambda x: int(x)):
            sname    = tag_dm[tag_dm["Lining_System_Code"]==code]["Lining_System_Short_Name"].iloc[0]
            base_pct = _sc_pct(baseline, tag, code)
            best_gain = 0; best_pos = i; best_new_pct = base_pct
            for new_pos in range(0, i):
                new_order = [t for t in tag_list if t != tag]
                new_order.insert(new_pos, tag)
                alloc   = cascade_allocate(new_order)
                new_pct = _sc_pct(alloc, tag, code)
                gain    = new_pct - base_pct
                if gain > best_gain:
                    best_gain = gain; best_pos = new_pos; best_new_pct = new_pct
            if best_gain > 0.4:
                sc_suggestions.append({
                    "tag": tag, "code": code, "sname": sname,
                    "current_pos": i+1, "suggest_pos": best_pos+1,
                    "current_pct": base_pct, "new_pct": best_new_pct,
                    "gain": round(best_gain, 1), "is_full": best_new_pct >= 99.9,
                })

    eq_suggestions.sort(key=lambda x: x["gain"], reverse=True)
    sc_suggestions.sort(key=lambda x: (x["is_full"], x["gain"]), reverse=True)
    return {
        "eq_suggestions": eq_suggestions[:5],
        "sc_suggestions": sc_suggestions[:8],
        "baseline_pcts":  baseline_pcts,
    }


def render_suggestion_panel(tag_list: list[str], panel_key: str) -> None:
    """Renders the Smart Reordering Suggestion panel."""
    st.markdown(
        '<div class="sec-hdr" style="margin-top:1.4rem;">'
        '💡 Smart Reordering Suggestions</div>', unsafe_allow_html=True)
    st.caption("Each suggestion shows the single best position change for one equipment "
               "or system code. Moving it there frees up inventory for that item first.")

    with st.spinner("Analysing reorder scenarios…"):
        result = _run_suggestion_engine(tag_list)

    eq_sugg = result["eq_suggestions"]
    sc_sugg = result["sc_suggestions"]

    if not eq_sugg and not sc_sugg:
        return  # hide silently when no improvements found

    c_eq, c_sc = st.columns(2, gap="large")

    with c_eq:
        st.markdown(
            '<div style="font-family:\'JetBrains Mono\',monospace;font-size:.65rem;'
            'font-weight:700;letter-spacing:.1em;text-transform:uppercase;'
            'color:var(--amber);margin-bottom:.6rem;">🔩 By Equipment</div>',
            unsafe_allow_html=True)
        if not eq_sugg:
            st.caption("No equipment-level gains found.")
        for s in eq_sugg:
            gc = "#10B981" if s["new_pct"] >= 99.9 else "#F59E0B"
            tn = eq_master.set_index("Equipment_Tag_No.")["Name"].get(s["tag"], s["tag"])
            tag_label = f"{s['tag']}  ·  {tn[:24]}"
            move_label = (
                f"Move #{s['current_pos']} → #{s['suggest_pos']}  ·  "
                f"{s['current_pct']:.1f}% → {s['new_pct']:.1f}%"
                + ("  ✅ Full completion!" if s['new_pct'] >= 99.9 else "")
            )
            st.markdown(
                f'<div class="card" style="margin-bottom:.45rem;padding:.75rem 1rem;">'
                f'<div style="display:flex;justify-content:space-between;align-items:center;">'
                f'<span style="font-size:.82rem;font-weight:700;color:var(--t0);">{tag_label}</span>'
                f'<span style="font-family:JetBrains Mono,monospace;font-size:.75rem;'
                f'color:{gc};font-weight:700;">+{s["gain"]:.1f}%</span></div>'
                f'<div style="font-size:.76rem;color:var(--t2);margin-top:.35rem;">'
                f'{move_label}</div></div>',
                unsafe_allow_html=True)

    with c_sc:
        st.markdown(
            '<div style="font-family:\'JetBrains Mono\',monospace;font-size:.65rem;'
            'font-weight:700;letter-spacing:.1em;text-transform:uppercase;'
            'color:var(--amber);margin-bottom:.6rem;">⚙️ By System Code</div>',
            unsafe_allow_html=True)
        if not sc_sugg:
            st.caption("No system-code-level gains found.")
        for s in sc_sugg:
            gc   = "#10B981" if s["is_full"] else "#F59E0B"
            full = ('  <span style="background:rgba(16,185,129,.15);color:#10B981;'
                    'font-size:.65rem;padding:.1rem .4rem;border-radius:3px;">100% COMPLETE</span>'
                    if s["is_full"] else "")
            tn = eq_master.set_index("Equipment_Tag_No.")["Name"].get(s["tag"], s["tag"])
            sc_label   = f"Code {s['code']}  {s['sname']}  ·  {s['tag']} {tn[:20]}"
            sc_move    = (
                f"Move #{s['current_pos']} → #{s['suggest_pos']}  ·  "
                f"{s['current_pct']:.1f}% → {s['new_pct']:.1f}%"
                + ("  ✅ 100% COMPLETE" if s["is_full"] else "")
            )
            st.markdown(
                f'<div class="card" style="margin-bottom:.45rem;padding:.75rem 1rem;">'
                f'<div style="display:flex;justify-content:space-between;align-items:center;">'
                f'<span style="font-size:.8rem;font-weight:600;color:var(--t0);">{sc_label}</span>'
                f'<span style="font-family:JetBrains Mono,monospace;font-size:.75rem;'
                f'color:{gc};font-weight:700;">+{s["gain"]:.1f}%</span></div>'
                f'<div style="font-size:.76rem;color:var(--t2);margin-top:.3rem;">'
                f'{sc_move}</div></div>',
                unsafe_allow_html=True)


# ── Type/Description label helper ─────────────────────────────────────────────
def _eq_label(tag: str) -> str:
    """Returns 'TAG  —  Name  |  Type  |  Desc' for display labels."""
    row = eq_master[eq_master["Equipment_Tag_No."] == tag]
    if row.empty: return tag
    r = row.iloc[0]
    type_s = str(r.get("Type","") or "").strip()
    desc_s = str(r.get("Description","") or "").strip()
    name_s = str(r.get("Name","") or "").strip()
    parts  = [name_s]
    if type_s and type_s not in ("nan","—"): parts.append(type_s)
    if desc_s and desc_s not in ("nan","—"): parts.append(desc_s[:28])
    return f"{tag}  —  " + "  |  ".join(parts)


def plotly_mat_table(df: pd.DataFrame, key_suffix: str, height: int = 380,
                     show_sqm: bool = False, tag: str = "", code: str = "") -> None:
    """Colour-coded material table. If show_sqm=True, adds SQM columns after qty cols."""
    base_cols = ["Material_Code", "Material_Name", "UOM",
                 "Demand_Qty", "Allocated_Qty", "Shortfall_Qty", "Fulfillment_Pct"]
    avail_cols = [c for c in base_cols if c in df.columns]
    df2 = df[avail_cols].copy()

    # Add Ordered_Qty if available
    if "Ordered_Qty" in df.columns:
        df2.insert(df2.columns.get_loc("Allocated_Qty"), "Ordered_Qty", df["Ordered_Qty"])

    rename_map = {
        "Material_Code":  "Code",
        "Material_Name":  "Material Name",
        "UOM":            "UOM",
        "Demand_Qty":     "Demand",
        "Ordered_Qty":    "On Order",
        "Allocated_Qty":  "Allocated",
        "Shortfall_Qty":  "Shortfall",
        "Fulfillment_Pct":"Fulfil %",
    }
    df2 = df2.rename(columns=rename_map)

    # SQM columns — per-material, based on each material's own fulfillment
    if show_sqm and tag and code:
        total_sqm_sc = sqm_ref[
            (sqm_ref["Equipment_Tag_No."]==tag) &
            (sqm_ref["Lining_System_Code"]==code)
        ]["Total_SQM"].sum()
        # Each material gets SQM proportional to its own fulfillment rate
        mat_fulfill = df["Fulfillment_Pct"].values / 100.0 if "Fulfillment_Pct" in df.columns else                       (df["Allocated_Qty"] / df["Demand_Qty"].replace(0, np.nan)).fillna(1.0).clip(0,1).values
        df2["SQM Total"]   = round(total_sqm_sc, 2)
        df2["SQM Done"]    = (total_sqm_sc * mat_fulfill).round(2)
        df2["SQM Deficit"] = (total_sqm_sc * (1 - mat_fulfill)).round(2)
        df2["SQM Done %"]  = (mat_fulfill * 100).round(1)

    # Pre-computed per-row SQM columns (for combined/aggregated tables)
    elif {"SQM_Total","SQM_Done","SQM_Deficit"}.issubset(df.columns):
        df2["SQM Total"]   = df["SQM_Total"].values
        df2["SQM Done"]    = df["SQM_Done"].values
        df2["SQM Deficit"] = df["SQM_Deficit"].values
        df2["SQM Done %"]  = np.where(
            df["SQM_Total"].values > 0,
            (df["SQM_Done"].values / df["SQM_Total"].replace(0, np.nan).values * 100),
            100.0
        ).round(1)

    fmt = {
        "Demand":     "{:,.3f}",
        "Allocated":  "{:,.3f}",
        "Shortfall":  "{:,.3f}",
        "Fulfil %":   "{:.1f}%",
    }
    if "On Order" in df2.columns:
        fmt["On Order"] = "{:,.3f}"
    if "SQM Total" in df2.columns:
        fmt.update({"SQM Total":"{:,.2f}","SQM Done":"{:,.2f}",
                    "SQM Deficit":"{:,.2f}","SQM Done %":"{:.1f}%"})

    fulfil_col = "Fulfil %"

    def style_row(row):
        pct = row.get(fulfil_col, 100)
        if pd.isna(pct): pct = 100.0
        if pct >= 100:
            bg, tc = "rgba(16,185,129,0.12)", "#10B981"
        elif pct >= 90:
            bg, tc = "rgba(249,115,22,0.12)",  "#F97316"
        elif pct >= 80:
            bg, tc = "rgba(234,179,8,0.12)",   "#EAB308"
        else:
            bg, tc = "rgba(239,68,68,0.12)",   "#EF4444"
        styles = [f"background-color:{bg}"] * len(row)
        fidx = list(row.index).index(fulfil_col) if fulfil_col in row.index else -1
        if fidx >= 0:
            styles[fidx] = f"background-color:{bg};color:{tc};font-weight:700"
        return styles

    styled_df = df2.style.apply(style_row, axis=1).format(fmt)
    st.dataframe(styled_df, hide_index=True, use_container_width=True,
                 height=height, key=f"tbl_{key_suffix}")



# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:.5rem 0 1.2rem">
      <div style="font-family:'JetBrains Mono',monospace;font-size:1rem;
                  font-weight:700;color:#F59E0B;">🏗 SME</div>
      <div style="font-family:'JetBrains Mono',monospace;font-size:.56rem;
                  letter-spacing:.18em;text-transform:uppercase;color:var(--t5);margin-top:2px;">
        Smart Material Estimator v3</div>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="sec-hdr">📍 Project Overview</div>', unsafe_allow_html=True)
    loc_counts = eq_master.groupby("Location")["Equipment_Tag_No."].count()
    for loc in LOCATION_ORDER:
        cnt   = loc_counts.get(loc, 0)
        badge = loc_badge(loc)
        st.markdown(
            f'<div style="display:flex;justify-content:space-between;'
            f'align-items:center;margin-bottom:.35rem;">'
            f'{badge}'
            f'<span style="font-family:\'JetBrains Mono\',monospace;'
            f'font-size:.75rem;color:var(--t3);">{cnt} equip.</span></div>',
            unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="sec-hdr">📦 Inventory</div>', unsafe_allow_html=True)
    st.caption(f"📦 {len(inv)} materials  ·  "
               f"⚠️ {(inv['Available_Qty']==0).sum()} at zero stock")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="sec-hdr">📋 Session</div>', unsafe_allow_html=True)
    n_sess = len(st.session_state.session_tags)
    if n_sess:
        for t in st.session_state.session_tags:
            st.caption(f"  · {t}")
        if st.button("🗑 Clear Session", key="clear_sidebar"):
            st.session_state.session_tags = []
            st.rerun()
    else:
        st.caption("No equipment added yet.")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""<div style="font-family:'JetBrains Mono',monospace;font-size:.6rem;
        letter-spacing:.08em;color:var(--t5);">
        🟢 100%  🟠 90–99%  🟡 80–89%  🔴 &lt;80%</div>""",
        unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# STICKY HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="sticky-header-wrap">
  <div style="display:flex;align-items:baseline;gap:.9rem;margin-bottom:.35rem;">
    <span style="font-family:'JetBrains Mono',monospace;font-size:1.2rem;
                 font-weight:700;color:var(--t0);letter-spacing:-.01em;">
      🏗 Smart Material Estimator</span>
    <span style="font-size:.72rem;color:var(--t4);letter-spacing:.03em;">
      System-code level · Cascading allocation · Priority-based</span>
  </div>
</div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────────────
tab0, tab1, tab2, tab3, tab4, tab_consume, tab5 = st.tabs([
    "📊  Dashboard",
    "🔍  Equipment Entry",
    "📦  Session Order Report",
    "📍  Location Report",
    "⚙️  Execution Plan",
    "📝  Daily Consumption",
    "📈  Total Overview",
])


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 0 · DASHBOARD (Project Overview + Material Requirement & Procurement)
# ═══════════════════════════════════════════════════════════════════════════════
with tab0:

    # ── Dashboard toggle ──────────────────────────────────────────────────────
    dash_view = st.radio(
        "View", ["📈 Project Overview", "🛒 Material Requirement & Procurement"],
        horizontal=True, key="dash_view", label_visibility="collapsed",
    )
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Shared filter controls (used by both views) ───────────────────────────
    st.markdown('<div class="sec-hdr">🎛 Filter</div>', unsafe_allow_html=True)
    df1, df2_col, df3, df4 = st.columns(4)
    with df1:
        sel_locations = st.multiselect("📍 Location", options=LOCATION_ORDER,
                                        default=LOCATION_ORDER, key="dash_loc")
    with df2_col:
        all_types_d = sorted(eq_master["Type"].str.strip().unique().tolist())
        sel_types = st.multiselect("🏷 Type", options=all_types_d,
                                    default=all_types_d, key="dash_type")
    with df3:
        all_codes_d = (
            dm[["Lining_System_Code","Lining_System_Short_Name"]].drop_duplicates()
            .sort_values("Lining_System_Code", key=lambda x: x.astype(int))
        )
        code_opts_d = [f"Code {r.Lining_System_Code} – {r.Lining_System_Short_Name}"
                       for _, r in all_codes_d.iterrows()]
        sel_codes_raw = st.multiselect("⚙️ System Code", options=code_opts_d,
                                        default=code_opts_d, key="dash_code")
        sel_codes = [c.split(" – ")[0].replace("Code ","").strip() for c in sel_codes_raw]
    with df4:
        all_desc_d = sorted(eq_master["Description"].dropna().unique().tolist())
        sel_desc = st.multiselect("📋 Description", options=all_desc_d,
                                   default=all_desc_d, key="dash_desc")

    # ── Apply filters ─────────────────────────────────────────────────────────
    filtered_eq = eq_master[
        eq_master["Location"].isin(sel_locations) &
        eq_master["Type"].str.strip().isin(sel_types) &
        eq_master["Description"].isin(sel_desc)
    ]
    if sel_codes:
        tags_w_code = dm[dm["Lining_System_Code"].isin(sel_codes)]["Equipment_Tag_No."].unique()
        filtered_eq = filtered_eq[filtered_eq["Equipment_Tag_No."].isin(tags_w_code)]
    filtered_tags = filtered_eq["Equipment_Tag_No."].tolist()

    filtered_dm = dm[
        (dm["Equipment_Tag_No."].isin(filtered_tags)) &
        (dm["Lining_System_Code"].isin(sel_codes))
    ].copy()

    # ── Correct SQM using sqm_ref (never sum from dm — inflated x n_materials) ─
    proj_sqm = sqm_ref[
        sqm_ref["Equipment_Tag_No."].isin(filtered_tags) &
        sqm_ref["Lining_System_Code"].isin(sel_codes)
    ]["Total_SQM"].sum()

    # ── Material demand aggregation ───────────────────────────────────────────
    f_demand = (
        filtered_dm.groupby(["Material_Code","Material_Name","UOM"], as_index=False)
        ["Demand_Qty"].sum()
    )
    f_demand = f_demand.merge(
        inv[["Material_Code","Available_Qty","Ordered_Qty"]], on="Material_Code", how="left"
    )
    f_demand["Available_Qty"] = f_demand["Available_Qty"].fillna(0)
    f_demand["Ordered_Qty"]   = f_demand["Ordered_Qty"].fillna(0)
    f_demand["Shortfall"]     = (f_demand["Demand_Qty"] - f_demand["Available_Qty"]).clip(lower=0).round(3)
    f_demand["Net_Shortfall"] = (
        f_demand["Demand_Qty"] - f_demand["Available_Qty"] - f_demand["Ordered_Qty"]
    ).clip(lower=0).round(3)
    f_demand["Coverage_Pct"]  = (
        f_demand["Available_Qty"].clip(upper=f_demand["Demand_Qty"]) /
        f_demand["Demand_Qty"].replace(0, np.nan) * 100
    ).fillna(100).clip(0,100).round(1)

    f_total_demand = f_demand["Demand_Qty"].sum()
    f_total_avail  = f_demand["Available_Qty"].clip(upper=f_demand["Demand_Qty"]).sum()
    f_total_short  = f_demand["Shortfall"].sum()
    f_total_net    = f_demand["Net_Shortfall"].sum()
    f_cov          = (f_total_avail / f_total_demand * 100) if f_total_demand > 0 else 100
    can_sqm        = round(proj_sqm * min(1.0, f_cov/100), 2)
    short_sqm      = round(proj_sqm - can_sqm, 2)

    # ─────────────────────────────────────────────────────────────────────────
    if dash_view == "📈 Project Overview":
    # ─────────────────────────────────────────────────────────────────────────

        # KPI strip
        k1,k2,k3,k4,k5,k6,k7 = st.columns(7)
        k1.metric("Equipment", len(filtered_tags),
                  help="Equipment tags matching current filter selection.")
        k2.metric("Total SQM", f"{proj_sqm:,.1f}",
                  help="Remaining surface area (m²) after deducting daily consumption entries.")
        k3.metric("SQM Achievable", f"{can_sqm:,.1f}  ({f_cov:.0f}%)",
                  help="SQM completable with current Available Qty = Total SQM × Coverage %.")
        k4.metric("SQM Deficit", f"{short_sqm:,.1f}",
                  help="SQM we cannot complete due to material shortfalls = Total − Achievable.")
        k5.metric("Overall Coverage", f"{f_cov:.1f}%", delta=f"{f_cov-100:.1f}%",
                  help="Allocated Qty ÷ Demand Qty × 100 across all filtered materials.")
        k6.metric("Total Shortfall", f"{f_total_short:,.0f}",
                  help="Sum of all material units short (Demand − Available, clipped at 0).")
        k7.metric("Critical (<50%)", int((f_demand["Coverage_Pct"]<50).sum()),
                  help="Materials where Available Qty covers less than 50% of total demand.")
        st.markdown("<br>", unsafe_allow_html=True)

        row1a, row1b = st.columns([1,1.6], gap="large")

        with row1a:
            st.markdown('<div class="sec-hdr">🎯 Overall Coverage</div>', unsafe_allow_html=True)
            fig_g = go.Figure(go.Indicator(
                mode="gauge+number+delta",
                value=round(f_cov,1),
                delta={"reference":100,"valueformat":".1f",
                       "decreasing":{"color":"#EF4444"},"increasing":{"color":"#10B981"}},
                number={"suffix":"%","font":{"family":"JetBrains Mono","size":36,"color":"var(--t0)"}},
                gauge={
                    "axis":{"range":[0,100],"tickwidth":1,
                            "tickfont":{"family":"JetBrains Mono","size":9}},
                    "bar":{"color":(
                        "#10B981" if f_cov>=100 else "#F97316" if f_cov>=90
                        else "#EAB308" if f_cov>=80 else "#EF4444"), "thickness":0.28},
                    "bgcolor":"rgba(0,0,0,0)","borderwidth":0,
                    "steps":[
                        {"range":[0,50],"color":"rgba(239,68,68,.08)"},
                        {"range":[50,80],"color":"rgba(234,179,8,.08)"},
                        {"range":[80,90],"color":"rgba(249,115,22,.08)"},
                        {"range":[90,100],"color":"rgba(16,185,129,.08)"},
                    ],
                },
                title={"text":f"Coverage  ·  {can_sqm:,.0f} / {proj_sqm:,.0f} SQM achievable",
                       "font":{"family":"JetBrains Mono","size":9,"color":"rgba(148,163,184,.7)"}},
            ))
            fig_g.update_layout(paper_bgcolor="rgba(0,0,0,0)",
                                margin=dict(l=20,r=20,t=30,b=10),height=240)
            st.plotly_chart(fig_g, use_container_width=True, key="dash_gauge")

            # Demand vs Available mini stacked bar
            fig_dm = go.Figure()
            fig_dm.add_trace(go.Bar(name="Available",x=["Inventory"],y=[f_total_avail],
                marker_color="#10B981",marker_opacity=.8,
                text=[f"{f_total_avail:,.0f}"],textposition="auto",
                textfont=dict(family="JetBrains Mono",size=10)))
            fig_dm.add_trace(go.Bar(name="Shortfall",x=["Inventory"],y=[f_total_short],
                marker_color="#EF4444",marker_opacity=.8,
                text=[f"{f_total_short:,.0f}"],textposition="auto",
                textfont=dict(family="JetBrains Mono",size=10)))
            fig_dm.update_layout(barmode="stack",paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",margin=dict(l=0,r=0,t=0,b=0),height=120,
                showlegend=True,
                legend=dict(orientation="h",y=1.15,x=0,
                            font=dict(family="JetBrains Mono",size=9),bgcolor="rgba(0,0,0,0)"),
                xaxis=dict(showgrid=False,showticklabels=False),
                yaxis=dict(gridcolor="rgba(128,128,128,.1)",showticklabels=False))
            st.plotly_chart(fig_dm, use_container_width=True, key="dash_dmini")

        with row1b:
            st.markdown('<div class="sec-hdr">📍 Coverage by Location (SQM)</div>',
                        unsafe_allow_html=True)
            loc_rows = []
            for loc in sel_locations:
                loc_tags = filtered_eq[filtered_eq["Location"]==loc]["Equipment_Tag_No."].tolist()
                loc_dm_f = filtered_dm[filtered_dm["Equipment_Tag_No."].isin(loc_tags)]
                if loc_dm_f.empty: continue
                loc_agg = (loc_dm_f.groupby("Material_Code",as_index=False)["Demand_Qty"].sum()
                    .merge(inv[["Material_Code","Available_Qty"]],on="Material_Code",how="left"))
                loc_agg["Available_Qty"] = loc_agg["Available_Qty"].fillna(0)
                loc_d = loc_agg["Demand_Qty"].sum()
                loc_a = loc_agg["Available_Qty"].clip(upper=loc_agg["Demand_Qty"]).sum()
                loc_s = (loc_agg["Demand_Qty"]-loc_agg["Available_Qty"]).clip(lower=0).sum()
                loc_c = (loc_a/loc_d*100) if loc_d>0 else 100
                # ✅ Correct SQM using sqm_ref
                loc_sqm = sqm_ref[
                    sqm_ref["Equipment_Tag_No."].isin(loc_tags) &
                    sqm_ref["Lining_System_Code"].isin(sel_codes)
                ]["Total_SQM"].sum()
                loc_can = round(loc_sqm * min(1.0, loc_c/100), 2)
                loc_rows.append({"Location":loc,"Equipment":len(loc_tags),
                    "SQM":loc_sqm,"SQM_Can":loc_can,"SQM_Short":round(loc_sqm-loc_can,2),
                    "Demand":loc_d,"Available":loc_a,"Shortfall":loc_s,"Coverage_%":round(loc_c,1)})

            if loc_rows:
                loc_df = pd.DataFrame(loc_rows)
                loc_colors_map = {"Brown Field":"#3B82F6","TRAIN J":"#F59E0B","TRAIN K":"#10B981"}
                fig_loc = go.Figure()
                for _, lr in loc_df.iterrows():
                    c = loc_colors_map.get(lr["Location"],"#94A3B8")
                    fig_loc.add_trace(go.Bar(
                        x=[lr["Location"]],y=[lr["SQM_Can"]],name=f'{lr["Location"]} Can Do',
                        marker_color=c,marker_opacity=.8,
                        text=[f'{lr["Coverage_%"]:.0f}%\n{lr["SQM_Can"]:,.0f} SQM'],
                        textposition="inside",textfont=dict(family="JetBrains Mono",size=10,color="#fff"),
                        showlegend=False))
                    fig_loc.add_trace(go.Bar(
                        x=[lr["Location"]],y=[lr["SQM_Short"]],name=f'{lr["Location"]} Deficit',
                        marker_color="#EF4444",marker_opacity=.6,
                        text=[f'{lr["SQM_Short"]:,.0f} SQM deficit'],
                        textposition="inside",textfont=dict(family="JetBrains Mono",size=9,color="#fff"),
                        showlegend=False))
                fig_loc.update_layout(barmode="stack",paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",margin=dict(l=0,r=0,t=10,b=0),height=220,
                    xaxis=dict(tickfont=dict(family="JetBrains Mono",size=11)),
                    yaxis=dict(gridcolor="rgba(128,128,128,.08)",
                               title=dict(text="SQM",font=dict(family="JetBrains Mono",size=9))),
                    font=dict(family="JetBrains Mono",size=10,color="rgba(148,163,184,.8)"))
                st.plotly_chart(fig_loc, use_container_width=True, key="dash_loc_bar")

                # Location stat cards
                cols_loc = st.columns(len(loc_rows))
                for col, lr in zip(cols_loc, loc_rows):
                    dot = ("🟢" if lr["Coverage_%"]>=100 else "🟠" if lr["Coverage_%"]>=90
                           else "🟡" if lr["Coverage_%"]>=80 else "🔴")
                    loc_html = (
                        '<div class="card" style="text-align:center;padding:.7rem;">'
                        f'<div style="font-size:1.1rem;">{dot}</div>'
                        '<div style="font-family:\'JetBrains Mono\',monospace;font-size:.72rem;'
                        f'font-weight:700;color:var(--amber);margin:.2rem 0;">{lr["Location"]}</div>'
                        '<div style="font-family:\'JetBrains Mono\',monospace;font-size:1.1rem;'
                        f'font-weight:700;color:var(--t0);">{lr["Coverage_%"]:.1f}%</div>'
                        f'<div style="font-size:.68rem;color:var(--t3);">'
                        f'{lr["SQM_Can"]:,.0f} / {lr["SQM"]:,.0f} SQM</div>'
                        f'<div style="font-size:.65rem;color:var(--t3);">{lr["Equipment"]} equipment</div>'
                        '</div>'
                    )
                    col.markdown(loc_html, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        row2a, row2b = st.columns(2, gap="large")

        with row2a:
            st.markdown('<div class="sec-hdr">⚙️ Coverage by System Code (SQM)</div>',
                        unsafe_allow_html=True)
            sc_rows = []
            for code in sel_codes:
                sc_dm = filtered_dm[filtered_dm["Lining_System_Code"]==code]
                if sc_dm.empty: continue
                sname = sc_dm["Lining_System_Short_Name"].iloc[0]
                sc_agg = (sc_dm.groupby("Material_Code",as_index=False)["Demand_Qty"].sum()
                    .merge(inv[["Material_Code","Available_Qty"]],on="Material_Code",how="left"))
                sc_agg["Available_Qty"] = sc_agg["Available_Qty"].fillna(0)
                sc_d = sc_agg["Demand_Qty"].sum()
                sc_a = sc_agg["Available_Qty"].clip(upper=sc_agg["Demand_Qty"]).sum()
                sc_s = (sc_agg["Demand_Qty"]-sc_agg["Available_Qty"]).clip(lower=0).sum()
                sc_c = (sc_a/sc_d*100) if sc_d>0 else 100
                # ✅ Correct SQM
                sc_sqm = sqm_ref[
                    sqm_ref["Equipment_Tag_No."].isin(filtered_tags) &
                    (sqm_ref["Lining_System_Code"]==code)
                ]["Total_SQM"].sum()
                sc_can = round(sc_sqm * min(1.0, sc_c/100), 2)
                sc_rows.append({"Code":f"Code {code}","Short_Name":sname,
                    "SQM":sc_sqm,"SQM_Can":sc_can,"SQM_Short":round(sc_sqm-sc_can,2),
                    "Coverage_%":round(sc_c,1)})

            if sc_rows:
                sc_df = pd.DataFrame(sc_rows).sort_values("Coverage_%")
                bar_c = ["#10B981" if c>=100 else "#F97316" if c>=90
                         else "#EAB308" if c>=80 else "#EF4444" for c in sc_df["Coverage_%"]]
                fig_sc = go.Figure(go.Bar(
                    y=sc_df["Code"]+"  "+sc_df["Short_Name"],x=sc_df["Coverage_%"],
                    orientation="h",marker_color=bar_c,marker_opacity=.8,
                    text=[f"{v:.0f}%  ({r['SQM_Can']:,.0f}/{r['SQM']:,.0f} SQM)"
                          for v,(_,r) in zip(sc_df["Coverage_%"],sc_df.iterrows())],
                    textposition="inside",textfont=dict(family="JetBrains Mono",size=9,color="#fff"),
                ))
                fig_sc.add_vline(x=100,line_color="rgba(128,128,128,.2)",
                                 line_dash="dot",line_width=1)
                fig_sc.update_layout(paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    margin=dict(l=0,r=30,t=5,b=0),
                    height=max(220,len(sc_df)*42),
                    xaxis=dict(range=[0,115],gridcolor="rgba(128,128,128,.08)",
                               tickfont=dict(family="JetBrains Mono",size=9)),
                    yaxis=dict(gridcolor="rgba(128,128,128,.08)",
                               tickfont=dict(family="JetBrains Mono",size=9)),
                    font=dict(family="JetBrains Mono",size=10,color="rgba(148,163,184,.8)"))
                st.plotly_chart(fig_sc, use_container_width=True, key="dash_sc_bar")

                sc_show = sc_df.copy()
                sc_show.columns = ["Code","Short Name","SQM Total","SQM Can Do",
                                   "SQM Deficit","Coverage %"]
                sc_show[["SQM Total","SQM Can Do","SQM Deficit"]] = (
                    sc_show[["SQM Total","SQM Can Do","SQM Deficit"]].round(1))
                st.dataframe(sc_show,use_container_width=True,hide_index=True,
                             key="dash_sc_tbl")

        with row2b:
            st.markdown('<div class="sec-hdr">🧪 Coverage by Material</div>',
                        unsafe_allow_html=True)
            mat_rows_d = f_demand.copy().sort_values("Coverage_Pct")
            if not mat_rows_d.empty:
                bar_cm = ["#10B981" if c>=100 else "#F97316" if c>=90
                          else "#EAB308" if c>=80 else "#EF4444"
                          for c in mat_rows_d["Coverage_Pct"]]
                fig_mat = go.Figure(go.Bar(
                    y=mat_rows_d["Material_Code"]+"  "+mat_rows_d["Material_Name"].fillna("").str[:18],
                    x=mat_rows_d["Coverage_Pct"],orientation="h",
                    marker_color=bar_cm,marker_opacity=.8,
                    text=[f"{v:.0f}%" for v in mat_rows_d["Coverage_Pct"]],
                    textposition="inside",textfont=dict(family="JetBrains Mono",size=9,color="#fff"),
                    customdata=mat_rows_d[["Available_Qty","Demand_Qty","Shortfall","UOM"]].values,
                    hovertemplate=(
                        "<b>%{y}</b><br>Coverage: %{x:.1f}%<br>"
                        "Available: %{customdata[0]:,.1f} %{customdata[3]}<br>"
                        "Demand: %{customdata[1]:,.1f} %{customdata[3]}<br>"
                        "Shortfall: %{customdata[2]:,.1f} %{customdata[3]}<extra></extra>"
                    ),
                ))
                fig_mat.add_vline(x=100,line_color="rgba(128,128,128,.2)",
                                  line_dash="dot",line_width=1)
                fig_mat.update_layout(paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",margin=dict(l=0,r=30,t=5,b=0),
                    height=max(340,len(mat_rows_d)*34),
                    xaxis=dict(range=[0,115],gridcolor="rgba(128,128,128,.08)",
                               tickfont=dict(family="JetBrains Mono",size=9)),
                    yaxis=dict(gridcolor="rgba(128,128,128,.08)",
                               tickfont=dict(family="JetBrains Mono",size=9)),
                    font=dict(family="JetBrains Mono",size=10,color="rgba(148,163,184,.8)"))
                st.plotly_chart(fig_mat, use_container_width=True, key="dash_mat_bar")

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="sec-hdr">📋 Full Material Balance</div>',
                    unsafe_allow_html=True)
        tbl_d = f_demand.sort_values("Coverage_Pct").copy()
        tbl_show = tbl_d[["Material_Code","Material_Name","UOM",
                           "Available_Qty","Ordered_Qty","Demand_Qty","Shortfall","Net_Shortfall","Coverage_Pct"]].copy()
        tbl_show.columns = ["Code","Material Name","UOM",
                             "Available","On Order","Total Demand","Shortfall","Net Shortfall","Coverage %"]

        def _style_cov(row):
            pct = row["Coverage %"]
            if pd.isna(pct): pct = 100.0
            if pct>=100:   bg,tc = "rgba(16,185,129,.1)","#10B981"
            elif pct>=90:  bg,tc = "rgba(249,115,22,.1)","#F97316"
            elif pct>=80:  bg,tc = "rgba(234,179,8,.1)", "#EAB308"
            else:          bg,tc = "rgba(239,68,68,.1)", "#EF4444"
            styles = [f"background-color:{bg}"]*len(row)
            styles[-1] = f"background-color:{bg};color:{tc};font-weight:700"
            return styles

        styled_tbl = tbl_show.style.apply(_style_cov,axis=1).format({
            "Available":"{:,.2f}","On Order":"{:,.2f}",
            "Total Demand":"{:,.2f}","Shortfall":"{:,.2f}",
            "Net Shortfall":"{:,.2f}","Coverage %":"{:.1f}%"})
        st.dataframe(styled_tbl,use_container_width=True,hide_index=True,
                     height=50+len(tbl_show)*35,key="dash_mat_tbl")

        # ── Stock-only materials (in inventory but not used in any recipe/demand) ──
        recipe_codes = set(dm["Material_Code"].unique())
        stock_only = inv[~inv["Material_Code"].isin(recipe_codes)].copy()
        if not stock_only.empty:
            st.markdown(
                '<div class="sec-hdr" style="margin-top:.8rem;">'
                '📦 Stock-Only Materials (No Demand in Any System Code)</div>',
                unsafe_allow_html=True)
            st.caption(
                "These materials are in your inventory but are not used "
                "in any lining system recipe. No demand is generated for them.")
            so_show = stock_only[["Material_Code","Material_Name","UOM",
                                   "Available_Qty","Ordered_Qty"]].copy()
            so_show["Ordered_Qty"] = so_show["Ordered_Qty"].fillna(0)
            so_show.columns = ["Code","Material Name","UOM","Available","On Order"]
            st.dataframe(so_show, use_container_width=True, hide_index=True,
                         key="dash_stock_only")

        da, db = st.columns(2)
        with da:
            st.download_button("⬇ Download Material Balance",
                data=excel_bytes(tbl_show.reset_index(drop=True)),
                file_name="dashboard_material_balance.xlsx",
                mime="application/vnd.ms-excel",use_container_width=True)

    # ─────────────────────────────────────────────────────────────────────────
    else:  # 🛒 Material Requirement & Procurement
    # ─────────────────────────────────────────────────────────────────────────

        st.markdown('<div class="sec-hdr">🛒 Material Requirement & Procurement — Location / System Code View</div>',
                    unsafe_allow_html=True)

        # KPI strip
        p1,p2,p3,p4,p5,p6 = st.columns(6)
        p1.metric("Equipment",        len(filtered_tags))
        p2.metric("Total SQM",        f"{proj_sqm:,.1f}")
        p3.metric("SQM Achievable",   f"{can_sqm:,.1f}  ({f_cov:.0f}%)")
        p4.metric("SQM Deficit",      f"{short_sqm:,.1f}")
        p5.metric("Shortfall Units",  f"{f_total_short:,.0f}")
        p6.metric("After Orders (Net)",f"{f_total_net:,.0f}")
        st.markdown("<br>", unsafe_allow_html=True)

        # Per-location, per-system-code breakdown
        for loc in sel_locations:
            loc_tags = filtered_eq[filtered_eq["Location"]==loc]["Equipment_Tag_No."].tolist()
            if not loc_tags: continue

            loc_dm = filtered_dm[filtered_dm["Equipment_Tag_No."].isin(loc_tags)]
            if loc_dm.empty: continue

            loc_sqm = sqm_ref[
                sqm_ref["Equipment_Tag_No."].isin(loc_tags) &
                sqm_ref["Lining_System_Code"].isin(sel_codes)
            ]["Total_SQM"].sum()

            loc_agg = (loc_dm.groupby("Material_Code",as_index=False)["Demand_Qty"].sum()
                .merge(inv[["Material_Code","Available_Qty","Ordered_Qty"]],
                       on="Material_Code",how="left"))
            loc_agg["Available_Qty"] = loc_agg["Available_Qty"].fillna(0)
            loc_agg["Ordered_Qty"]   = loc_agg["Ordered_Qty"].fillna(0)
            loc_agg["Shortfall"]     = (loc_agg["Demand_Qty"]-loc_agg["Available_Qty"]).clip(lower=0)
            loc_agg["Net_Shortfall"] = (loc_agg["Demand_Qty"]-loc_agg["Available_Qty"]-loc_agg["Ordered_Qty"]).clip(lower=0)
            loc_d = loc_agg["Demand_Qty"].sum()
            loc_a = loc_agg["Available_Qty"].clip(upper=loc_agg["Demand_Qty"]).sum()
            loc_c = (loc_a/loc_d*100) if loc_d>0 else 100
            loc_can_sqm = round(loc_sqm * min(1.0, loc_c/100), 2)
            loc_sh_sqm  = round(loc_sqm - loc_can_sqm, 2)
            loc_dot = "🟢" if loc_c>=100 else "🟠" if loc_c>=90 else "🟡" if loc_c>=80 else "🔴"

            badge_cls = {"Brown Field":"loc-bf","TRAIN J":"loc-tj","TRAIN K":"loc-tk"}.get(loc,"loc-bf")
            st.markdown(
                f'<div style="display:flex;align-items:center;gap:.8rem;margin:1.2rem 0 .5rem;">'
                f'<span style="font-size:.95rem;">{loc_dot}</span>'
                f'<span class="loc-badge {badge_cls}">{loc}</span>'
                f'<span style="font-family:\'JetBrains Mono\',monospace;font-size:.75rem;'
                f'color:var(--t3);">{len(loc_tags)} equip  ·  {loc_can_sqm:,.1f}/{loc_sqm:,.1f} SQM  ·  {loc_c:.1f}%</span>'
                f'</div>', unsafe_allow_html=True)

            # Per system code within this location
            for code in sorted(sel_codes, key=lambda x: int(x)):
                code_dm = loc_dm[loc_dm["Lining_System_Code"]==code]
                if code_dm.empty: continue
                sname = code_dm["Lining_System_Short_Name"].iloc[0]

                # ✅ Correct SQM
                code_sqm = sqm_ref[
                    sqm_ref["Equipment_Tag_No."].isin(loc_tags) &
                    (sqm_ref["Lining_System_Code"]==code)
                ]["Total_SQM"].sum()

                code_agg = (code_dm.groupby(["Material_Code","Material_Name","UOM"],
                                            as_index=False)["Demand_Qty"].sum()
                    .merge(inv[["Material_Code","Available_Qty","Ordered_Qty"]],
                           on="Material_Code",how="left"))
                code_agg["Available_Qty"] = code_agg["Available_Qty"].fillna(0)
                code_agg["Ordered_Qty"]   = code_agg["Ordered_Qty"].fillna(0)
                code_agg["Shortfall"]     = (code_agg["Demand_Qty"]-code_agg["Available_Qty"]).clip(lower=0).round(3)
                code_agg["Net_Shortfall"] = (code_agg["Demand_Qty"]-code_agg["Available_Qty"]-code_agg["Ordered_Qty"]).clip(lower=0).round(3)
                code_agg["Coverage_Pct"]  = (
                    code_agg["Available_Qty"].clip(upper=code_agg["Demand_Qty"]) /
                    code_agg["Demand_Qty"].replace(0,np.nan)*100
                ).fillna(100).clip(0,100).round(1)
                code_agg["Fulfillment_Pct"] = code_agg["Coverage_Pct"]

                cd = code_agg["Demand_Qty"].sum()
                ca = code_agg["Available_Qty"].clip(upper=code_agg["Demand_Qty"]).sum()
                cc = (ca/cd*100) if cd>0 else 100
                c_can_sqm = round(code_sqm * min(1.0, cc/100), 2)
                c_sh_sqm  = round(code_sqm - c_can_sqm, 2)
                c_dot = "🟢" if cc>=100 else "🟠" if cc>=90 else "🟡" if cc>=80 else "🔴"

                with st.expander(
                    f"{c_dot}  Code {code} – {sname}  ·  "
                    f"{c_can_sqm:,.1f}/{code_sqm:,.1f} SQM  ·  {cc:.1f}%",
                    expanded=False,
                ):
                    pc1,pc2,pc3,pc4,pc5 = st.columns(5)
                    pc1.metric("System Code", f"Code {code}")
                    pc2.metric("Short Name",  sname)
                    pc3.metric("SQM Total",   f"{code_sqm:,.2f}")
                    pc4.metric("SQM Achievable", f"{c_can_sqm:,.2f}")
                    pc5.metric("SQM Deficit",    f"{c_sh_sqm:,.2f}")

                    # Table with Available, On Order, Demand, Shortfall, Net Shortfall
                    tbl_proc = code_agg[["Material_Code","Material_Name","UOM",
                                         "Available_Qty","Ordered_Qty","Demand_Qty",
                                         "Shortfall","Net_Shortfall","Fulfillment_Pct"]].copy()
                    tbl_proc.columns = ["Code","Material Name","UOM",
                                        "Available","On Order","Demand",
                                        "Shortfall","Net Shortfall (After Orders)","Fulfil %"]

                    def _style_proc(row):
                        pct = row["Fulfil %"]
                        if pd.isna(pct): pct = 100.0
                        if pct>=100:  bg,tc = "rgba(16,185,129,.1)","#10B981"
                        elif pct>=90: bg,tc = "rgba(249,115,22,.1)","#F97316"
                        elif pct>=80: bg,tc = "rgba(234,179,8,.1)", "#EAB308"
                        else:         bg,tc = "rgba(239,68,68,.1)", "#EF4444"
                        styles = [f"background-color:{bg}"]*len(row)
                        styles[-1] = f"background-color:{bg};color:{tc};font-weight:700"
                        return styles

                    styled_proc = tbl_proc.style.apply(_style_proc,axis=1).format({
                        "Available":"{:,.3f}","On Order":"{:,.3f}","Demand":"{:,.3f}",
                        "Shortfall":"{:,.3f}","Net Shortfall (After Orders)":"{:,.3f}",
                        "Fulfil %":"{:.1f}%"})
                    st.dataframe(styled_proc,use_container_width=True,hide_index=True,
                                 height=65+len(tbl_proc)*35,
                                 key=f"proc_{loc}_{code}")

            st.markdown('<div style="border-bottom:1px solid var(--border);margin:.8rem 0;"></div>',
                        unsafe_allow_html=True)

        # Grand total procurement table
        st.markdown('<div class="sec-hdr" style="margin-top:1rem;">📦 Grand Total — All Selected Equipment</div>',
                    unsafe_allow_html=True)

        grand = f_demand.sort_values("Coverage_Pct").copy()
        grand_show = grand[["Material_Code","Material_Name","UOM",
                             "Available_Qty","Ordered_Qty","Demand_Qty",
                             "Shortfall","Net_Shortfall","Coverage_Pct"]].copy()
        grand_show.columns = ["Code","Material Name","UOM",
                               "Available","On Order","Demand",
                               "Shortfall","Net Shortfall","Coverage %"]

        def _style_grand(row):
            pct = row["Coverage %"]
            if pd.isna(pct): pct=100.0
            if pct>=100:  bg,tc = "rgba(16,185,129,.1)","#10B981"
            elif pct>=90: bg,tc = "rgba(249,115,22,.1)","#F97316"
            elif pct>=80: bg,tc = "rgba(234,179,8,.1)", "#EAB308"
            else:         bg,tc = "rgba(239,68,68,.1)", "#EF4444"
            styles = [f"background-color:{bg}"]*len(row)
            styles[-1] = f"background-color:{bg};color:{tc};font-weight:700"
            return styles

        styled_grand = grand_show.style.apply(_style_grand,axis=1).format({
            "Available":"{:,.2f}","On Order":"{:,.2f}","Demand":"{:,.2f}",
            "Shortfall":"{:,.2f}","Net Shortfall":"{:,.2f}","Coverage %":"{:.1f}%"})
        st.dataframe(styled_grand,use_container_width=True,hide_index=True,
                     height=50+len(grand_show)*35,key="proc_grand_tbl")

        gc1, gc2 = st.columns(2)
        with gc1:
            st.download_button("⬇ Download Grand Procurement Table",
                data=excel_bytes(grand_show.reset_index(drop=True)),
                file_name="procurement_grand_total.xlsx",
                mime="application/vnd.ms-excel",use_container_width=True)
        with gc2:
            shortage_net = grand[grand["Net_Shortfall"]>0][
                ["Material_Code","Material_Name","UOM","Available_Qty","Ordered_Qty",
                 "Demand_Qty","Shortfall","Net_Shortfall"]].copy()
            shortage_net.columns = ["Code","Material Name","UOM","Available","On Order",
                                    "Demand","Shortfall","NET TO ORDER"]
            if not shortage_net.empty:
                st.download_button("⬇ Net Order List Only",
                    data=excel_bytes(shortage_net.reset_index(drop=True)),
                    file_name="net_order_list.xlsx",
                    mime="application/vnd.ms-excel",use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 · EQUIPMENT ENTRY
# ═══════════════════════════════════════════════════════════════════════════════
with tab1:
    left, right = st.columns([1, 1.65], gap="large")

    # ── LEFT: filter + search + session list ────────────────────────────────
    with left:
        st.markdown('<div class="sec-hdr">🎛 Filter Equipment</div>',
                    unsafe_allow_html=True)

        # 3 filter selectors
        f_loc  = st.multiselect("📍 Location", options=LOCATION_ORDER,
                                 default=[], key="t1_loc",
                                 placeholder="All locations")
        all_types = sorted(eq_master["Type"].str.strip().unique().tolist())
        f_type = st.multiselect("🏷 Type", options=all_types,
                                 default=[], key="t1_type",
                                 placeholder="All types")
        all_codes_t1 = sorted(
            dm["Lining_System_Code"].unique().tolist(), key=lambda x: int(x))
        f_code = st.multiselect(
            "⚙️ System Code", options=all_codes_t1,
            format_func=lambda c: f"Code {c} – "
                f"{dm[dm['Lining_System_Code']==c]['Lining_System_Short_Name'].iloc[0]}",
            default=[], key="t1_code", placeholder="All system codes")

        # Build filtered tag list
        filtered_eq = eq_master.copy()
        if f_loc:
            filtered_eq = filtered_eq[filtered_eq["Location"].isin(f_loc)]
        if f_type:
            filtered_eq = filtered_eq[filtered_eq["Type"].str.strip().isin(f_type)]
        if f_code:
            tags_with_code = dm[dm["Lining_System_Code"].isin(f_code)][
                "Equipment_Tag_No."].unique().tolist()
            filtered_eq = filtered_eq[filtered_eq["Equipment_Tag_No."].isin(tags_with_code)]
        filtered_tags_t1 = sorted(filtered_eq["Equipment_Tag_No."].tolist())

        st.markdown('<div class="sec-hdr" style="margin-top:.8rem;">🔍 Find Equipment</div>',
                    unsafe_allow_html=True)
        selected_tag = st.selectbox(
            "tag_search", [""] + filtered_tags_t1,
            format_func=lambda t: (
                "" if t == "" else
                f"{t}  —  "
                f"{eq_master.set_index('Equipment_Tag_No.')['Name'].get(t,'')}"
            ),
            key="tag_select", label_visibility="collapsed",
        )
        ca, cb = st.columns([2,1])
        with ca:
            already = selected_tag in st.session_state.session_tags
            add_btn = st.button("＋ Add to Session", key="add_btn",
                                disabled=(selected_tag=="" or already))
        with cb:
            if already and selected_tag:
                st.markdown(
                    '<div style="padding:.45rem 0;font-family:\'JetBrains Mono\','
                    'monospace;font-size:.7rem;color:#10B981;">✓ In session</div>',
                    unsafe_allow_html=True)
        if add_btn and selected_tag:
            st.session_state.session_tags.append(selected_tag)
            st.rerun()

        # ── Priority order controls ──────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="sec-hdr">📋 Session Priority List</div>',
                    unsafe_allow_html=True)

        session_tags = st.session_state.session_tags
        if not session_tags:
            st.info("Add equipment tags above to build your session.")
        else:
            from streamlit_sortables import sort_items

            alloc_df = cascade_allocate(session_tags)
            tag_name = eq_master.set_index("Equipment_Tag_No.")["Name"].to_dict()
            tag_loc  = eq_master.set_index("Equipment_Tag_No.")["Location"].to_dict()

            # Static labels — no % to keep component state stable across reruns
            def _item_label(i, t):
                name = tag_name.get(t, t)[:28]
                loc  = tag_loc.get(t, "—")
                return f"#{i+1}  ||  {t}  ||  {name}  ||  {loc}"

            display_items = [_item_label(i, t) for i, t in enumerate(session_tags)]

            # Key tied to exact list — forces full re-init when items added/removed
            sort_key = "sess_sort_" + "_".join(session_tags)
            st.caption("⠿ Drag to reorder — order is applied instantly.")
            sorted_display = sort_items(display_items, direction="vertical", key=sort_key)

            # Parse tag back using explicit || delimiter
            def _parse_tag(label):
                parts = label.split("  ||  ")
                return parts[1].strip() if len(parts) > 1 else label.strip()

            new_order = [_parse_tag(l) for l in sorted_display if _parse_tag(l) in session_tags]
            # Safety fallback
            if len(new_order) != len(session_tags):
                new_order = session_tags[:]

            # Auto-apply if order changed
            if new_order != st.session_state.session_tags:
                st.session_state.session_tags = new_order
                st.rerun()

            # Show fulfillment rows with ✕ remove button
            alloc_df2 = cascade_allocate(new_order)
            for idx_t, t in enumerate(new_order):
                pct  = tag_fulfillment(alloc_df2, t)
                name = tag_name.get(t, t)
                loc  = tag_loc.get(t, "—")
                dot  = status_dot(pct)
                tag_total_sqm = sqm_ref[sqm_ref["Equipment_Tag_No."]==t]["Total_SQM"].sum()
                can_sqm  = round(tag_total_sqm * min(1.0, pct/100), 2)
                row_c, row_x = st.columns([9, 1])
                with row_c:
                    # Append system codes to session list display
                    _t_codes = sorted(
                        dm[dm["Equipment_Tag_No."]==t]["Lining_System_Code"].unique(),
                        key=lambda x: int(x))
                    _codes_badges = "  ".join(
                        f'<span style="font-family:\'JetBrains Mono\',monospace;'
                        f'font-size:.58rem;background:rgba(245,158,11,.15);'
                        f'color:var(--amber);border-radius:3px;padding:.1rem .3rem;">C{c}</span>'
                        for c in _t_codes)
                    _sess_parts = [
                        f'<div class="session-equip" style="margin-bottom:.22rem;">',
                        f'<span class="{dot}" style="font-family:JetBrains Mono,monospace;',
                        f'font-size:.75rem;font-weight:600;color:var(--t1);">{t}</span>',
                        f'<span style="font-size:.75rem;color:var(--t3);margin-left:.5rem;">{name[:22]}</span>',
                        f'<span style="margin-left:.4rem;">{_codes_badges}</span>',
                        f'<span style="font-family:JetBrains Mono,monospace;',
                        f'font-size:.62rem;color:var(--t4);margin-left:.4rem;">{loc}</span>',
                        f'<span style="float:right;font-family:JetBrains Mono,monospace;',
                        f'font-size:.64rem;color:var(--t3);">{can_sqm:,.1f}/{tag_total_sqm:,.1f} SQM&nbsp;&nbsp;</span>',
                        f'<span style="float:right;">{fulfil_pill(pct)}</span>',
                        '</div>',
                    ]
                    st.markdown("".join(_sess_parts), unsafe_allow_html=True)
                with row_x:
                    if st.button("✕", key=f"rm_{t}_{idx_t}", help=f"Remove {t}"):
                        st.session_state.session_tags.remove(t)
                        st.rerun()

            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑 Clear All", key="clear_all"):
                st.session_state.session_tags = []
                st.rerun()

    # ── RIGHT: equipment info card + system-code material tables ─────────────
    with right:
        if not selected_tag:
            st.markdown("""
            <div style="text-align:center;padding:4rem 1rem;">
              <div style="font-size:2.5rem;opacity:.12;margin-bottom:.8rem;">🔩</div>
              <div style="font-family:'JetBrains Mono',monospace;font-size:.75rem;
                          color:var(--t5);letter-spacing:.1em;">
                SELECT AN EQUIPMENT TAG TO VIEW DETAILS</div></div>""",
                unsafe_allow_html=True)
        else:
            row = eq_master[eq_master["Equipment_Tag_No."]==selected_tag].iloc[0]
            tag_codes = dm[dm["Equipment_Tag_No."]==selected_tag][
                ["Lining_System_Code","Lining_System_Short_Name","Total_SQM"]
            ].drop_duplicates().sort_values("Lining_System_Code", key=lambda x: x.astype(int))

            # Info card
            st.markdown(
                f'<div class="card card-amber">'
                f'<div style="display:flex;justify-content:space-between;'
                f'align-items:flex-start;margin-bottom:.7rem;">'
                f'<div><div style="font-family:\'JetBrains Mono\',monospace;'
                f'font-size:.62rem;color:var(--t4);letter-spacing:.1em;'
                f'text-transform:uppercase;">Equipment Tag</div>'
                f'<div style="font-family:\'JetBrains Mono\',monospace;'
                f'font-size:1.1rem;font-weight:700;color:#F59E0B;">'
                f'{row["Equipment_Tag_No."]}</div></div>'
                f'{loc_badge(str(row["Location"]))}</div>'
                f'<div style="font-size:.95rem;font-weight:600;color:var(--t0);'
                f'margin-bottom:.9rem;">{row["Name"]}</div>'
                f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr;'
                f'gap:.4rem .7rem;font-size:.78rem;">'
                f'<div><span style="color:var(--t4);">Type</span><br>'
                f'<span style="color:var(--t1);">{row["Type"]}</span></div>'
                f'<div><span style="color:var(--t4);">Description</span><br>'
                f'<span style="color:var(--t1);">{row["Description"] or "—"}</span></div>'
                f'<div><span style="color:var(--t4);">Material Spec.</span><br>'
                f'<span style="color:var(--t1);">{row["Material_Spec"] or "—"}</span></div>'
                f'<div style="grid-column:1/-1;">'
                f'<span style="color:var(--t4);">Lining Systems</span><br>'
                f'<span style="color:var(--t2);font-size:.75rem;line-height:1.5;">'
                f'{str(row["Lining_Systems"]).replace(chr(10),"<br>")}'
                f'</span></div></div></div>',
                unsafe_allow_html=True)

            # System code sections
            st.markdown('<div class="sec-hdr" style="margin-top:1rem;">'
                        '⚗️ System Code Material Requirements</div>',
                        unsafe_allow_html=True)

            for _, sc_row in tag_codes.iterrows():
                code  = sc_row["Lining_System_Code"]
                sname = sc_row["Lining_System_Short_Name"]
                sqm   = sc_row["Total_SQM"]

                mat_rows = dm[
                    (dm["Equipment_Tag_No."]==selected_tag) &
                    (dm["Lining_System_Code"]==code)
                ].copy()
                mat_rows = mat_rows.merge(
                    inv[["Material_Code","Available_Qty"]], on="Material_Code", how="left")
                mat_rows["Available_Qty"]   = mat_rows["Available_Qty"].fillna(0)
                mat_rows["Allocated_Qty"]   = mat_rows["Available_Qty"].clip(
                    upper=mat_rows["Demand_Qty"])
                mat_rows["Shortfall_Qty"]   = (
                    mat_rows["Demand_Qty"] - mat_rows["Allocated_Qty"]).clip(lower=0)
                mat_rows["Fulfillment_Pct"] = (
                    mat_rows["Allocated_Qty"] /
                    mat_rows["Demand_Qty"].replace(0,np.nan) * 100
                ).fillna(100).clip(0,100).round(2)

                d_sum = mat_rows["Demand_Qty"].sum()
                a_sum = mat_rows["Allocated_Qty"].sum()
                pct   = min(100, a_sum/d_sum*100) if d_sum > 0 else 100
                short = mat_rows["Shortfall_Qty"].sum()

                with st.expander(
                    f"System Code {code}  ·  {sname}  ·  {sqm:,.2f} SQM  "
                    f"·  Coverage: {pct:.1f}%",
                    expanded=False,
                ):
                    mi1,mi2,mi3,mi4 = st.columns(4)
                    mi1.metric("System Code",   code)
                    mi2.metric("Short Name",    sname)
                    mi3.metric("Surface Area",  f"{sqm:,.2f} SQM")
                    mi4.metric("Coverage",      f"{pct:.1f}%")
                    plotly_mat_table(
                        mat_rows, f"entry_{selected_tag}_{code}",
                        height=65 + len(mat_rows)*30
                    )

            # Grand total for this equipment
            all_mat = dm[dm["Equipment_Tag_No."]==selected_tag].merge(
                inv[["Material_Code","Available_Qty"]], on="Material_Code", how="left")
            all_mat["Available_Qty"] = all_mat["Available_Qty"].fillna(0)
            all_mat["Shortfall"]     = (
                all_mat["Demand_Qty"] -
                all_mat["Available_Qty"].clip(upper=all_mat["Demand_Qty"])
            ).clip(lower=0)
            gt_demand = all_mat["Demand_Qty"].sum()
            gt_alloc  = all_mat["Available_Qty"].clip(upper=all_mat["Demand_Qty"]).sum()
            gt_short  = all_mat["Shortfall"].sum()
            gt_pct    = min(100, gt_alloc/gt_demand*100) if gt_demand > 0 else 100

            st.markdown(f"""
            <div class="grand-box" style="margin-top:.8rem;">
              <div style="font-family:'JetBrains Mono',monospace;font-size:.6rem;
                          letter-spacing:.14em;text-transform:uppercase;
                          color:#F59E0B;margin-bottom:.7rem;">
                Equipment Grand Total — {selected_tag}</div>
              <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:.7rem;">
                <div><div style="font-family:'JetBrains Mono',monospace;
                     font-size:1.3rem;font-weight:700;color:var(--t0);">
                     {len(tag_codes)}</div>
                     <div style="font-size:.62rem;text-transform:uppercase;
                     letter-spacing:.08em;color:var(--t4);">System Codes</div></div>
                <div><div style="font-family:'JetBrains Mono',monospace;
                     font-size:1.3rem;font-weight:700;color:var(--t0);">
                     {gt_demand:,.0f}</div>
                     <div style="font-size:.62rem;text-transform:uppercase;
                     letter-spacing:.08em;color:var(--t4);">Total Demand</div></div>
                <div><div style="font-family:'JetBrains Mono',monospace;
                     font-size:1.3rem;font-weight:700;
                     color:{'#EF4444' if gt_short>0 else '#10B981'};">
                     {gt_short:,.1f}</div>
                     <div style="font-size:.62rem;text-transform:uppercase;
                     letter-spacing:.08em;color:var(--t4);">Total Shortfall</div></div>
                <div><div style="font-family:'JetBrains Mono',monospace;
                     font-size:1.3rem;font-weight:700;
                     color:{'#10B981' if gt_pct>=100 else '#F97316' if gt_pct>=90 else '#EAB308' if gt_pct>=80 else '#EF4444'};">
                     {gt_pct:.1f}%</div>
                     <div style="font-size:.62rem;text-transform:uppercase;
                     letter-spacing:.08em;color:var(--t4);">Coverage</div></div>
              </div>
            </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 · SESSION ORDER REPORT
# ═══════════════════════════════════════════════════════════════════════════════
with tab2:
    session_tags = st.session_state.session_tags

    if not session_tags:
        st.info("Add equipment tags in the Entry tab to generate a report.")
    else:
        alloc_df = cascade_allocate(session_tags)
        tag_name = eq_master.set_index("Equipment_Tag_No.")["Name"].to_dict()
        tag_loc  = eq_master.set_index("Equipment_Tag_No.")["Location"].to_dict()

        # Session KPIs
        tot_demand = alloc_df["Demand_Qty"].sum()
        tot_alloc  = alloc_df["Allocated_Qty"].sum()
        tot_short  = alloc_df["Shortfall_Qty"].sum()
        ov_pct     = min(100, tot_alloc/tot_demand*100) if tot_demand > 0 else 100
        n_mats     = alloc_df["Material_Code"].nunique()
        n_short_m  = alloc_df[alloc_df["Shortfall_Qty"]>0]["Material_Code"].nunique()

        k1,k2,k3,k4,k5 = st.columns(5)
        k1.metric("Equipment",       len(session_tags))
        k2.metric("Materials",       n_mats)
        k3.metric("Need to Order",   n_short_m)
        k4.metric("Total Shortfall", f"{tot_short:,.1f}")
        k5.metric("Overall Coverage",f"{ov_pct:.1f}%")
        st.markdown("<br>", unsafe_allow_html=True)

        # ── Priority reorder (updates global session_tags) ───────────────────
        st.markdown('<div class="sec-hdr">⠿ Drag to Reorder Priority — changes reflect everywhere</div>',
                    unsafe_allow_html=True)

        from streamlit_sortables import sort_items as _sort2

        def _t2_sl(i, t):
            return f"#{i+1}  ||  {t}  ||  {tag_name.get(t,t)[:28]}  ||  {tag_loc.get(t,'—')}"

        t2_display = [_t2_sl(i, t) for i, t in enumerate(session_tags)]
        t2_key     = "t2_sort_" + "_".join(session_tags)
        t2_sorted  = _sort2(t2_display, direction="vertical", key=t2_key)

        def _t2_parse(label):
            parts = label.split("  ||  ")
            return parts[1].strip() if len(parts) > 1 else label.strip()

        t2_new_order = [_t2_parse(l) for l in t2_sorted if _t2_parse(l) in session_tags]
        
        if len(t2_new_order) == len(session_tags) and t2_new_order != st.session_state.session_tags:
            st.session_state.session_tags = t2_new_order
            st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
        # ── Per-equipment expanders ──────────────────────────────────────────
        st.markdown('<div class="sec-hdr">Per-Equipment System Code Breakdown</div>',
                    unsafe_allow_html=True)

        for i, tag in enumerate(session_tags):
            tag_alloc = alloc_df[alloc_df["Equipment_Tag_No."]==tag]
            t_pct     = tag_fulfillment(alloc_df, tag)
            dot       = status_dot(t_pct)
            t_short   = tag_alloc["Shortfall_Qty"].sum()
            name      = tag_name.get(tag, tag)
            loc       = tag_loc.get(tag, "—")
            eq_row    = eq_master[eq_master["Equipment_Tag_No."]==tag].iloc[0]

            t_sqm      = sqm_ref[sqm_ref["Equipment_Tag_No."]==tag]["Total_SQM"].sum()
            t_can_sqm  = round(t_sqm * min(1.0, t_pct/100), 2)
            t_dot_char = "🟢" if t_pct>=100 else "🟠" if t_pct>=90 else "🟡" if t_pct>=80 else "🔴"
            _t2_type = str(eq_row.get("Type","") or "").strip()
            _t2_desc = str(eq_row.get("Description","") or "").strip()[:20]
            _t2_meta = "  |  ".join(p for p in [_t2_type,_t2_desc] if p and p not in ("nan","—"))
            with st.expander(
                f"{t_dot_char}  #{i+1}  {tag}  ·  {name}  ·  {_t2_meta}  ·  {loc}  "
                f"·  {t_can_sqm:,.1f}/{t_sqm:,.1f} SQM  ·  {t_pct:.1f}%",
                expanded=False,
            ):
                # Equipment meta strip
                m1,m2,m3,m4 = st.columns(4)
                m1.markdown(f'**Type:** {eq_row["Type"]}')
                m2.markdown(f'**Description:** {eq_row["Description"] or "—"}')
                m3.markdown(f'**Material Spec.:** {eq_row["Material_Spec"] or "—"}')
                m4.markdown(f'**Total SQM:** `{eq_row["Total_SQM"]:,.2f}`')
                st.caption(f'**Lining Systems:** '
                           f'{str(eq_row["Lining_Systems"]).replace(chr(10)," | ")}')
                st.markdown("---")

                # Per system code tables
                for code in sorted(tag_alloc["Lining_System_Code"].unique(),
                                   key=lambda x: int(x)):
                    code_alloc = tag_alloc[tag_alloc["Lining_System_Code"]==code].copy()
                    sname = code_alloc["Lining_System_Short_Name"].iloc[0]
                    sqm   = code_alloc["Total_SQM"].iloc[0]
                    c_pct = syscode_fulfillment(alloc_df, tag, code)
                    c_short = code_alloc["Shortfall_Qty"].sum()
                    c_demand = code_alloc["Demand_Qty"].sum()
                    c_alloc  = code_alloc["Allocated_Qty"].sum()

                    _, c_can_sqm, c_short_sqm = sqm_can_do(alloc_df, tag, code)
                    c_dot = "🟢" if c_pct>=100 else "🟠" if c_pct>=90 else "🟡" if c_pct>=80 else "🔴"
                    st.markdown(
                        f'<div class="syscode-block">'
                        f'<div class="syscode-hdr">'
                        f'<span style="font-size:.85rem;">{c_dot}</span>'
                        f'<span class="code-badge">Code {code}</span>'
                        f'<span style="font-size:.8rem;color:var(--t1);font-weight:500;">'
                        f'{sname}</span>'
                        f'<span style="font-family:\'JetBrains Mono\',monospace;'
                        f'font-size:.72rem;color:var(--t3);">'
                        f'{c_can_sqm:,.1f}/{sqm:,.1f} SQM</span>'
                        f'<span style="margin-left:auto;">{fulfil_pill(c_pct)}</span>'
                        f'</div></div>',
                        unsafe_allow_html=True)

                    sc1,sc2,sc3,sc4 = st.columns(4)
                    sc1.metric("Demand",    f"{c_demand:,.2f}")
                    sc2.metric("Allocated", f"{c_alloc:,.2f}")
                    sc3.metric("Shortfall", f"{c_short:,.2f}")
                    sc4.metric("SQM Deficit", f"{c_short_sqm:,.2f}")
                    plotly_mat_table(
                        code_alloc,
                        f"rep_{tag}_{code}",
                        height=65 + len(code_alloc)*30,
                        show_sqm=True, tag=tag, code=code
                    )

                # Equipment grand total row
                st.markdown(
                    f'<div style="background:rgba(245,158,11,.05);'
                    f'border:1px solid rgba(245,158,11,.2);border-radius:6px;'
                    f'padding:.7rem 1rem;margin-top:.5rem;'
                    f'font-family:\'JetBrains Mono\',monospace;font-size:.8rem;">'
                    f'<span style="color:#F59E0B;font-weight:700;">GRAND TOTAL — {tag}</span>'
                    f'<span style="color:var(--t3);margin-left:1.5rem;">'
                    f'Demand: <strong style="color:var(--t1);">'
                    f'{tag_alloc["Demand_Qty"].sum():,.2f}</strong></span>'
                    f'<span style="color:var(--t3);margin-left:1rem;">'
                    f'Allocated: <strong style="color:var(--t1);">'
                    f'{tag_alloc["Allocated_Qty"].sum():,.2f}</strong></span>'
                    f'<span style="color:var(--t3);margin-left:1rem;">'
                    f'Shortfall: <strong style="color:#EF4444;">'
                    f'{tag_alloc["Shortfall_Qty"].sum():,.2f}</strong></span>'
                    f'<span style="margin-left:1rem;">{fulfil_pill(t_pct)}</span>'
                    f'</div>',
                    unsafe_allow_html=True)

        # ── Combined procurement list ─────────────────────────────────────────
        st.markdown('<div class="sec-hdr" style="margin-top:1.5rem;">'
                    '🛒 Combined Procurement List</div>',
                    unsafe_allow_html=True)

        combined = alloc_df.groupby(
            ["Material_Code","Material_Name","UOM"], as_index=False
        ).agg(
            Demand_Qty    =("Demand_Qty",    "sum"),
            Allocated_Qty =("Allocated_Qty", "sum"),
            Shortfall_Qty =("Shortfall_Qty", "sum"),
        )
        combined["Fulfillment_Pct"] = (
            combined["Allocated_Qty"] /
            combined["Demand_Qty"].replace(0, np.nan) * 100
        ).fillna(100).clip(0,100).round(2)

        # ── SQM per material: sum of SQM across every (tag,code) cell it touches,
        #    weighted by that cell's fulfillment ──────────────────────────────
        _sqm_per_mat = alloc_df.copy()
        _sqm_per_mat["SQM_Done_Cell"] = (
            _sqm_per_mat["Total_SQM"] * _sqm_per_mat["Fulfillment_Pct"] / 100
        )
        _sqm_agg = _sqm_per_mat.groupby("Material_Code", as_index=False).agg(
            SQM_Total =("Total_SQM",     "sum"),
            SQM_Done  =("SQM_Done_Cell", "sum"),
        )
        _sqm_agg["SQM_Deficit"] = (_sqm_agg["SQM_Total"] - _sqm_agg["SQM_Done"]).round(2)
        _sqm_agg["SQM_Total"]   = _sqm_agg["SQM_Total"].round(2)
        _sqm_agg["SQM_Done"]    = _sqm_agg["SQM_Done"].round(2)
        combined = combined.merge(_sqm_agg, on="Material_Code", how="left")
        combined = combined.sort_values("Fulfillment_Pct")

        # Stacked bar (shortage only)
        shortage_only = combined[combined["Shortfall_Qty"]>0].copy()
        if not shortage_only.empty:
            shortage_only["Label"] = (
                shortage_only["Material_Code"] + "  " +
                shortage_only["Material_Name"].fillna("").str[:22]
            )
            fig_bar = go.Figure()
            fig_bar.add_trace(go.Bar(
                name="Available", y=shortage_only["Label"],
                x=shortage_only["Allocated_Qty"], orientation="h",
                marker_color="#10B981", marker_opacity=.75,
                text=shortage_only["Allocated_Qty"].apply(lambda v:f"{v:,.0f}"),
                textposition="inside",
                textfont=dict(family="JetBrains Mono",size=9,color="#fff"),
            ))
            fig_bar.add_trace(go.Bar(
                name="To Order", y=shortage_only["Label"],
                x=shortage_only["Shortfall_Qty"], orientation="h",
                marker_color="#EF4444", marker_opacity=.75,
                text=shortage_only["Shortfall_Qty"].apply(lambda v:f"{v:,.1f}"),
                textposition="inside",
                textfont=dict(family="JetBrains Mono",size=9,color="#fff"),
            ))
            fig_bar.update_layout(
                barmode="stack",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(family="JetBrains Mono", size=10), # Fixed Font
                legend=dict(orientation="h", yanchor="bottom", y=1.02,
                            bgcolor="rgba(0,0,0,0)", font=dict(size=10)),
                margin=dict(l=0, r=60, t=28, b=0),
                height=max(300, len(shortage_only)*42),
                xaxis=dict(gridcolor="rgba(128, 128, 128, 0.2)", zerolinecolor="rgba(128, 128, 128, 0.2)"), # Fixed Grids
                yaxis=dict(gridcolor="rgba(128, 128, 128, 0.2)"), # Fixed Grids
            )
            st.plotly_chart(fig_bar, use_container_width=True, key="session_bar")

        plotly_mat_table(combined, "session_combined",
                         height=90+len(combined)*30)

        # Grand total box
        st.markdown(f"""
        <div class="grand-box" style="margin-top:1rem;">
          <div style="font-family:'JetBrains Mono',monospace;font-size:.6rem;
                      letter-spacing:.14em;text-transform:uppercase;
                      color:#F59E0B;margin-bottom:.7rem;">
            ⭐ Grand Total — {len(session_tags)} Equipment Session</div>
          <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:.7rem;">
            <div><div style="font-family:'JetBrains Mono',monospace;font-size:1.25rem;
                 font-weight:700;color:var(--t0);">{len(session_tags)}</div>
                 <div style="font-size:.6rem;text-transform:uppercase;
                 letter-spacing:.08em;color:var(--t4);">Equipment</div></div>
            <div><div style="font-family:'JetBrains Mono',monospace;font-size:1.25rem;
                 font-weight:700;color:var(--t0);">{n_mats}</div>
                 <div style="font-size:.6rem;text-transform:uppercase;
                 letter-spacing:.08em;color:var(--t4);">Materials</div></div>
            <div><div style="font-family:'JetBrains Mono',monospace;font-size:1.25rem;
                 font-weight:700;color:var(--t0);">{tot_demand:,.0f}</div>
                 <div style="font-size:.6rem;text-transform:uppercase;
                 letter-spacing:.08em;color:var(--t4);">Total Demand</div></div>
            <div><div style="font-family:'JetBrains Mono',monospace;font-size:1.25rem;
                 font-weight:700;color:#EF4444;">{n_short_m}</div>
                 <div style="font-size:.6rem;text-transform:uppercase;
                 letter-spacing:.08em;color:var(--t4);">To Procure</div></div>
            <div><div style="font-family:'JetBrains Mono',monospace;font-size:1.25rem;
                 font-weight:700;color:#EF4444;">{tot_short:,.1f}</div>
                 <div style="font-size:.6rem;text-transform:uppercase;
                 letter-spacing:.08em;color:var(--t4);">Shortfall Units</div></div>
          </div>
        </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        d1, d2 = st.columns(2)
        with d1:
            st.download_button("⬇ Full Session Report",
                               data=excel_bytes(alloc_df),
                               file_name="session_full_report.xlsx",
                               mime="application/vnd.ms-excel",
                               use_container_width=True)
        with d2:
            if not shortage_only.empty:
                st.download_button("⬇ Order List Only",
                                   data=excel_bytes(shortage_only),
                                   file_name="order_list.xlsx",
                                   mime="application/vnd.ms-excel",
                                   use_container_width=True)


        # ── Smart Reordering Suggestions ──────────────────────────────────────
        render_suggestion_panel(session_tags, "tab2")

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 · LOCATION REPORT
# ═══════════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="sec-hdr">📍 All Equipment by Location — Cascading Balance</div>',
                unsafe_allow_html=True)
    st.caption("Drag to reorder equipment within each location. Order is independent from the session list.")

    # ── Per-location order state (independent from global session_tags) ──────
    if "loc_order" not in st.session_state:
        st.session_state.loc_order = {}

    from streamlit_sortables import sort_items as _sort3

    loc_color = {
        "Brown Field": ("loc-bf","#3B82F6"),
        "TRAIN J":     ("loc-tj","#F59E0B"),
        "TRAIN K":     ("loc-tk","#10B981"),
    }

    for loc in LOCATION_ORDER:
        # Initialise from file order if not set
        default_loc_order = eq_master[eq_master["Location"]==loc]["Equipment_Tag_No."].tolist()
        if loc not in st.session_state.loc_order:
            st.session_state.loc_order[loc] = default_loc_order

        loc_tags_all = st.session_state.loc_order[loc]
        if not loc_tags_all:
            continue

        badge_cls, accent = loc_color.get(loc, ("loc-bf","#3B82F6"))

        # Compute quick fulfillment with CURRENT order for labels only
        loc_alloc_preview = cascade_allocate(loc_tags_all)

        # ── Location title above sortable list ──────────────────────────────
        st.markdown(
            f'<div style="font-family:\'JetBrains Mono\',monospace;font-size:.68rem;'
            f'font-weight:700;letter-spacing:.1em;color:var(--amber);margin-bottom:.3rem;">'
            f'📍 {loc} — Drag to set build priority</div>',
            unsafe_allow_html=True)

        # Static labels — no % to keep component state stable across reruns
        def _l3_label(i, t):
            name = eq_master.set_index("Equipment_Tag_No.")["Name"].get(t, t)[:26]
            sqm  = eq_master.set_index("Equipment_Tag_No.")["Total_SQM"].get(t, 0)
            return f"#{i+1}  ||  {t}  ||  {name}  ||  {sqm:,.1f} SQM"

        loc_display = [_l3_label(i, t) for i, t in enumerate(loc_tags_all)]
        loc_key     = f"loc_sort_{loc}_" + "_".join(loc_tags_all)
        loc_sorted  = _sort3(loc_display, direction="vertical", key=loc_key)

        def _l3_parse(label):
            parts = label.split("  ||  ")
            return parts[1].strip() if len(parts) > 1 else label.strip()

        new_loc_order = [_l3_parse(l) for l in loc_sorted if _l3_parse(l) in loc_tags_all]
        
        if len(new_loc_order) == len(loc_tags_all) and new_loc_order != st.session_state.loc_order[loc]:
            st.session_state.loc_order[loc] = new_loc_order
            st.rerun()

        # Re-cascade with (possibly new) order
        loc_alloc = cascade_allocate(loc_tags_all)

        loc_demand = loc_alloc["Demand_Qty"].sum()
        loc_short  = loc_alloc["Shortfall_Qty"].sum()
        loc_pct    = min(100,
            loc_alloc["Allocated_Qty"].sum()/loc_demand*100
        ) if loc_demand > 0 else 100

        # Location SQM (correct: sum unique (tag,code) SQM, not from dm)
        loc_sqm_total = sqm_ref[sqm_ref["Equipment_Tag_No."].isin(loc_tags_all)]["Total_SQM"].sum()
        loc_can_sqm   = round(loc_sqm_total * min(1.0, loc_pct/100), 2)
        loc_dot       = "🟢" if loc_pct>=100 else "🟠" if loc_pct>=90 else "🟡" if loc_pct>=80 else "🔴"

        # Location header
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:1rem;'
            f'margin:1.2rem 0 .7rem;">'
            f'<span style="font-size:.95rem;">{loc_dot}</span>'
            f'<span class="loc-badge {badge_cls}" style="font-size:.76rem;'
            f'padding:.28rem .9rem;">{loc}</span>'
            f'<span style="font-family:\'JetBrains Mono\',monospace;font-size:.75rem;'
            f'color:var(--t4);">{len(loc_tags_all)} equip</span>'
            f'<span style="font-family:\'JetBrains Mono\',monospace;font-size:.75rem;'
            f'color:var(--t3);">{loc_can_sqm:,.1f}/{loc_sqm_total:,.1f} SQM</span>'
            f'<span style="font-family:\'JetBrains Mono\',monospace;font-size:.75rem;'
            f'color:{"#EF4444" if loc_short>0 else "#10B981"};">'
            f'Shortfall: {loc_short:,.1f}</span>'
            f'<span style="font-family:\'JetBrains Mono\',monospace;font-size:.75rem;'
            f'color:var(--t3);">Coverage: {loc_pct:.1f}%</span>'
            f'</div>',
            unsafe_allow_html=True)

        # Per-equipment expanders
        for tag in loc_tags_all:
            tag_alloc = loc_alloc[loc_alloc["Equipment_Tag_No."]==tag]
            t_pct     = tag_fulfillment(loc_alloc, tag)
            t_short   = tag_alloc["Shortfall_Qty"].sum()
            eq_row    = eq_master[eq_master["Equipment_Tag_No."]==tag].iloc[0]
            dot       = status_dot(t_pct)

            _dot_char  = "🟢" if t_pct>=100 else "🟠" if t_pct>=90 else "🟡" if t_pct>=80 else "🔴"
            _t3_sqm    = sqm_ref[sqm_ref["Equipment_Tag_No."]==tag]["Total_SQM"].sum()
            _t3_cansqm = round(_t3_sqm * min(1.0, t_pct/100), 2)
            _t3_type   = str(eq_row.get("Type","") or "").strip()
            _t3_desc   = str(eq_row.get("Description","") or "").strip()[:20]
            _t3_meta   = "  |  ".join(p for p in [_t3_type,_t3_desc] if p and p not in ("nan","—"))
            with st.expander(
                f"{_dot_char}  {tag}  ·  {eq_row['Name']}  ·  {_t3_meta}  ·  "
                f"{_t3_cansqm:,.1f}/{_t3_sqm:,.1f} SQM  ·  {t_pct:.1f}%",
                expanded=False,
            ):
                c1,c2,c3 = st.columns(3)
                c1.markdown(f'**Type:** {eq_row["Type"]}')
                c2.markdown(f'**Description:** {eq_row["Description"] or "—"}')
                c3.markdown(f'**Material Spec.:** {eq_row["Material_Spec"] or "—"}')
                st.caption(
                    f'**Lining:** {str(eq_row["Lining_Systems"]).replace(chr(10)," | ")}')
                st.markdown("---")

                # Per system code
                for code in sorted(tag_alloc["Lining_System_Code"].unique(),
                                   key=lambda x: int(x)):
                    code_alloc = tag_alloc[tag_alloc["Lining_System_Code"]==code].copy()
                    sname = code_alloc["Lining_System_Short_Name"].iloc[0]
                    sqm   = code_alloc["Total_SQM"].iloc[0]
                    c_pct = syscode_fulfillment(loc_alloc, tag, code)

                    _, c3_can, c3_short_sqm = sqm_can_do(loc_alloc, tag, code)
                    c3_dot = "🟢" if c_pct>=100 else "🟠" if c_pct>=90 else "🟡" if c_pct>=80 else "🔴"
                    st.markdown(
                        f'<div class="syscode-block">'
                        f'<div class="syscode-hdr">'
                        f'<span style="font-size:.85rem;">{c3_dot}</span>'
                        f'<span class="code-badge">Code {code}</span>'
                        f'<span style="font-size:.8rem;color:var(--t1);">{sname}</span>'
                        f'<span style="font-family:\'JetBrains Mono\',monospace;'
                        f'font-size:.72rem;color:var(--t3);">'
                        f'{c3_can:,.1f}/{sqm:,.1f} SQM</span>'
                        f'<span style="margin-left:auto;">{fulfil_pill(c_pct)}</span>'
                        f'</div></div>',
                        unsafe_allow_html=True)
                    plotly_mat_table(
                        code_alloc,
                        f"loc_{loc}_{tag}_{code}",
                        height=65 + len(code_alloc)*30,
                        show_sqm=True, tag=tag, code=code
                    )

                # Equipment grand total
                st.markdown(
                    f'<div style="background:rgba(245,158,11,.05);'
                    f'border:1px solid rgba(245,158,11,.18);border-radius:6px;'
                    f'padding:.65rem .9rem;margin-top:.4rem;'
                    f'font-family:\'JetBrains Mono\',monospace;font-size:.77rem;">'
                    f'<span style="color:#F59E0B;font-weight:700;">TOTAL — {tag}</span>'
                    f'<span style="color:var(--t3);margin-left:1.2rem;">'
                    f'Demand: <b style="color:var(--t1);">'
                    f'{tag_alloc["Demand_Qty"].sum():,.2f}</b></span>'
                    f'<span style="color:var(--t3);margin-left:.8rem;">'
                    f'Shortfall: <b style="color:#EF4444;">'
                    f'{tag_alloc["Shortfall_Qty"].sum():,.2f}</b></span>'
                    f'<span style="margin-left:.8rem;">{fulfil_pill(t_pct)}</span>'
                    f'</div>',
                    unsafe_allow_html=True)

                # Add to session button
                if tag in st.session_state.session_tags:
                    st.markdown(
                        '<span style="font-family:\'JetBrains Mono\',monospace;'
                        'font-size:.7rem;color:#10B981;">✓ In session</span>',
                        unsafe_allow_html=True)
                else:
                    if st.button(f"＋ Add {tag} to Session",
                                 key=f"locadd_{loc}_{tag}"):
                        st.session_state.session_tags.append(tag)
                        st.rerun()

        # ── Bar chart (collapsible) per location ──────────────────────────────
        with st.expander(f"📊 Show Shortfall Chart — {loc}", expanded=False):
            chart_data = []
            for tag in loc_tags_all:
                tag_alloc = loc_alloc[loc_alloc["Equipment_Tag_No."]==tag]
                for code in sorted(tag_alloc["Lining_System_Code"].unique(),
                                   key=lambda x: int(x)):
                    ca = tag_alloc[tag_alloc["Lining_System_Code"]==code]
                    sname = ca["Lining_System_Short_Name"].iloc[0]
                    chart_data.append({
                        "Label": f"{tag}\nCode {code} ({sname})",
                        "Demand":    ca["Demand_Qty"].sum(),
                        "Allocated": ca["Allocated_Qty"].sum(),
                        "Shortfall": ca["Shortfall_Qty"].sum(),
                    })
            cdf = pd.DataFrame(chart_data)
            if not cdf.empty:
                fig_loc = go.Figure()
                fig_loc.add_trace(go.Bar(
                    name="Allocated", y=cdf["Label"], x=cdf["Allocated"],
                    orientation="h", marker_color=accent, marker_opacity=.65,
                    text=cdf["Allocated"].apply(lambda v:f"{v:,.0f}"),
                    textposition="inside",
                    textfont=dict(family="JetBrains Mono",size=8,color="#fff"),
                ))
                fig_loc.add_trace(go.Bar(
                    name="Shortfall", y=cdf["Label"], x=cdf["Shortfall"],
                    orientation="h", marker_color="#EF4444", marker_opacity=.75,
                    text=cdf["Shortfall"].apply(
                        lambda v:f"{v:,.0f}" if v > 0 else ""),
                    textposition="inside",
                    textfont=dict(family="JetBrains Mono",size=8,color="#fff"),
                ))
                fig_loc.update_layout(
                    barmode="stack",
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="JetBrains Mono",size=9,color="var(--t3)"),
                    legend=dict(orientation="h",yanchor="bottom",y=1.01,
                                bgcolor="rgba(0,0,0,0)"),
                    margin=dict(l=0,r=60,t=28,b=0),
                    height=max(350, len(cdf)*36),
                    xaxis=dict(gridcolor="#1E2E46",zerolinecolor="#1E2E46"),
                    yaxis=dict(gridcolor="#1E2E46"),
                    title=dict(text=f"{loc} — Demand vs Shortfall by System Code",
                               font=dict(family="JetBrains Mono",size=11,color="var(--t2)"))
                )
                st.plotly_chart(fig_loc, use_container_width=True,
                                key=f"loc_chart_{loc}")

        st.markdown(
            f'<div style="border-bottom:1px solid #1E2E46;margin:1rem 0;"></div>',
            unsafe_allow_html=True)


    # ── Per-location Excel downloads ────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="sec-hdr">📥 Download Report per Location</div>',
                unsafe_allow_html=True)

    dl_loc_cols = st.columns(len(LOCATION_ORDER))
    for _dl_col, _loc_dl in zip(dl_loc_cols, LOCATION_ORDER):
        _loc_tags_dl = st.session_state.loc_order.get(
            _loc_dl,
            eq_master[eq_master["Location"] == _loc_dl]["Equipment_Tag_No."].tolist()
        )
        if not _loc_tags_dl:
            continue
        _loc_alloc_dl = cascade_allocate(_loc_tags_dl)
        if _loc_alloc_dl.empty:
            continue
        # Enrich with inventory data for export
        _inv_lu = inv[["Material_Code","Available_Qty","Ordered_Qty"]].groupby(
            "Material_Code", as_index=False).first()
        _loc_report = _loc_alloc_dl.merge(_inv_lu, on="Material_Code", how="left")
        _loc_report["Available_Qty"] = _loc_report["Available_Qty"].fillna(0)
        _loc_report["Ordered_Qty"]   = _loc_report["Ordered_Qty"].fillna(0)
        _export_cols = [
            "Equipment_Tag_No.", "Lining_System_Code", "Lining_System_Short_Name",
            "Total_SQM", "Material_Code", "Material_Name", "UOM",
            "Demand_Qty", "Available_Qty", "Ordered_Qty",
            "Allocated_Qty", "Shortfall_Qty", "Fulfillment_Pct",
        ]
        _export_df = _loc_report[[c for c in _export_cols if c in _loc_report.columns]]
        _dl_col.download_button(
            label=f"⬇ {_loc_dl}",
            data=excel_bytes(_export_df),
            file_name=f"location_report_{_loc_dl.replace(' ', '_')}.xlsx",
            mime="application/vnd.ms-excel",
            use_container_width=True,
            key=f"dl_loc_{_loc_dl}",
        )

    # ── Print Report button ──────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="sec-hdr">🖨 Print Location Report</div>',
                unsafe_allow_html=True)
    st.markdown("""
    <button onclick="window.print()"
        style="font-family:'JetBrains Mono',monospace;font-size:.68rem;
               font-weight:700;letter-spacing:.08em;text-transform:uppercase;
               background:#F59E0B;color:#000;border:none;border-radius:4px;
               padding:.52rem 1.3rem;cursor:pointer;transition:all .15s;">
        🖨 Print / Save as PDF
    </button>
    <style>
    @media print {
        [data-testid="stSidebar"], [data-testid="stHeader"],
        .sticky-header-wrap, [data-testid="stTabs"] > div:first-of-type,
        button[onclick="window.print()"] { display:none!important; }
        [data-testid="stExpander"] { break-inside:avoid; }
        body { background:#fff!important; color:#000!important; }
    }
    </style>""", unsafe_allow_html=True)

    # ── Smart Reordering Suggestions per location ───────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    for _loc_sugg in LOCATION_ORDER:
        _loc_tags_sugg = st.session_state.loc_order.get(
            _loc_sugg,
            eq_master[eq_master["Location"]==_loc_sugg]["Equipment_Tag_No."].tolist()
        )
        if len(_loc_tags_sugg) < 2:
            continue
        with st.expander(f"💡 Smart Reordering Suggestions — {_loc_sugg}", expanded=False):
            render_suggestion_panel(_loc_tags_sugg, f"tab3_{_loc_sugg}")

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 · EXECUTION PLAN
# ═══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown('<div class="sec-hdr">⚙️ Execution Plan — Critical System Code Analysis</div>',
                unsafe_allow_html=True)

    session_tags = st.session_state.session_tags

    if not session_tags:
        st.info("Add equipment tags in the Entry tab first.")
    else:
        alloc_df   = cascade_allocate(session_tags)
        tag_name   = eq_master.set_index("Equipment_Tag_No.")["Name"].to_dict()

        sel_tag = st.selectbox(
            "Select Equipment",
            options=session_tags,
            format_func=lambda t: f"{t}  —  {tag_name.get(t,t)}",
            key="exec_tag",
        )

        tag_alloc = alloc_df[alloc_df["Equipment_Tag_No."]==sel_tag]
        avail_codes = sorted(tag_alloc["Lining_System_Code"].unique(),
                             key=lambda x: int(x))

        if not avail_codes:
            st.warning("No system code data for this equipment.")
        else:
            sel_code = st.selectbox(
                "Select Critical System Code",
                options=avail_codes,
                format_func=lambda c: (
                    f"Code {c}  —  "
                    f"{tag_alloc[tag_alloc['Lining_System_Code']==c]['Lining_System_Short_Name'].iloc[0]}"
                ),
                key="exec_code",
            )

            st.markdown("<br>", unsafe_allow_html=True)

            # Critical system code data
            crit = tag_alloc[tag_alloc["Lining_System_Code"]==sel_code].copy()
            crit_sname  = crit["Lining_System_Short_Name"].iloc[0]
            crit_sqm    = crit["Total_SQM"].iloc[0]
            crit_demand = crit["Demand_Qty"].sum()
            crit_alloc  = crit["Allocated_Qty"].sum()
            crit_short  = crit["Shortfall_Qty"].sum()
            crit_pct    = min(100, crit_alloc/crit_demand*100) if crit_demand > 0 else 100

            # Other system codes
            other_codes = [c for c in avail_codes if c != sel_code]
            other_alloc = tag_alloc[tag_alloc["Lining_System_Code"].isin(other_codes)]
            other_short = other_alloc["Shortfall_Qty"].sum()

            # ── Critical system code card ─────────────────────────────────────
            st.markdown(
                f'<div class="card card-amber">'
                f'<div style="font-family:\'JetBrains Mono\',monospace;'
                f'font-size:.6rem;letter-spacing:.14em;text-transform:uppercase;'
                f'color:#F59E0B;margin-bottom:.6rem;">Critical System Code</div>'
                f'<div style="display:flex;align-items:center;gap:1rem;'
                f'margin-bottom:.8rem;">'
                f'<span class="code-badge" style="font-size:.85rem;'
                f'padding:.3rem .8rem;">Code {sel_code}</span>'
                f'<span style="font-size:.95rem;font-weight:600;color:var(--t0);">'
                f'{crit_sname}</span>'
                f'<span style="font-family:\'JetBrains Mono\',monospace;'
                f'font-size:.8rem;color:var(--t3);">{crit_sqm:,.2f} SQM</span>'
                f'<span style="margin-left:auto;font-family:\'JetBrains Mono\','
                f'monospace;font-size:1.4rem;font-weight:700;'
                f'color:{"#10B981" if crit_pct>=100 else "#F97316" if crit_pct>=90 else "#EAB308" if crit_pct>=80 else "#EF4444"};">'
                f'{crit_pct:.1f}%</span>'
                f'</div>'
                f'<div style="font-size:.82rem;color:var(--t2);line-height:1.7;">'
                f'With current inventory allocation, <strong style="color:#F59E0B;">'
                f'{crit_pct:.1f}%</strong> of System Code {sel_code} ({crit_sname}) '
                f'can be completed for <strong style="color:var(--t0);">{sel_tag}</strong>. '
                + (
                    "✅ All materials for this system code are fully covered."
                    if crit_short == 0 else
                    f'⚠️ <strong style="color:#EF4444;">{crit_short:,.2f} units</strong>'
                    f' short across {(crit["Shortfall_Qty"]>0).sum()} material(s) — order these first to proceed.'
                ) +
                f'</div></div>',
                unsafe_allow_html=True)

            # Full materials table — all system codes
            st.markdown('<div class="sec-hdr" style="margin-top:1rem;">'
                        'All Materials — Status Overview</div>',
                        unsafe_allow_html=True)
            st.caption(
                f"Showing all {len(tag_alloc)} material rows across "
                f"{len(avail_codes)} system code(s) for {sel_tag}. "
                f"Cascade balance applied (priority position #{session_tags.index(sel_tag)+1})."
            )
            plotly_mat_table(
                tag_alloc.copy(),
                f"exec_all_{sel_tag}",
                height=80 + len(tag_alloc)*30
            )

            # ── ORDER PRIORITY SECTION ─────────────────────────────────────────
            st.markdown('<div class="sec-hdr" style="margin-top:1.2rem;">'
                        '📋 Procurement Order Priority</div>',
                        unsafe_allow_html=True)

            # 1️⃣ Critical code shortages
            crit_short_df = crit[crit["Shortfall_Qty"]>0][
                ["Material_Code","Material_Name","UOM","Demand_Qty",
                 "Allocated_Qty","Shortfall_Qty","Fulfillment_Pct"]
            ].copy()

            st.markdown(
                f'<div style="background:var(--red-bg);border:1px solid '
                f'rgba(239,68,68,.25);border-left:4px solid #EF4444;'
                f'border-radius:6px;padding:.8rem 1rem;margin-bottom:.6rem;">'
                f'<div style="font-family:\'JetBrains Mono\',monospace;font-size:.65rem;'
                f'letter-spacing:.12em;text-transform:uppercase;color:#EF4444;'
                f'margin-bottom:.4rem;">1️⃣ Order First — System Code {sel_code} '
                f'({crit_sname}) · Critical Path</div>'
                f'<div style="font-size:.8rem;color:var(--t2);">'
                f'{"No shortages on critical system code — fully covered ✅" if crit_short_df.empty else f"{len(crit_short_df)} material(s) need to be procured for this system code before work can begin."}'
                f'</div></div>',
                unsafe_allow_html=True)

            if not crit_short_df.empty:
                plotly_mat_table(crit_short_df,
                                 f"exec_crit_{sel_tag}_{sel_code}",
                                 height=65+len(crit_short_df)*30)

            # 2️⃣ Other system codes shortages
            for code in other_codes:
                code_alloc_df = tag_alloc[tag_alloc["Lining_System_Code"]==code]
                code_short    = code_alloc_df[code_alloc_df["Shortfall_Qty"]>0]
                sname_o       = code_alloc_df["Lining_System_Short_Name"].iloc[0]
                code_pct      = syscode_fulfillment(alloc_df, sel_tag, code)

                order_num = other_codes.index(code) + 2
                st.markdown(
                    f'<div style="background:var(--amber-bg);border:1px solid '
                    f'rgba(245,158,11,.2);border-left:4px solid #F59E0B;'
                    f'border-radius:6px;padding:.8rem 1rem;margin-bottom:.5rem;">'
                    f'<div style="font-family:\'JetBrains Mono\',monospace;'
                    f'font-size:.65rem;letter-spacing:.12em;text-transform:uppercase;'
                    f'color:#F59E0B;margin-bottom:.4rem;">'
                    f'{order_num}️⃣ Order Next — System Code {code} ({sname_o}) '
                    f'· Coverage: {code_pct:.1f}%</div>'
                    f'<div style="font-size:.8rem;color:var(--t2);">'
                    f'{"All materials covered ✅" if code_short.empty else f"{len(code_short)} material(s) short. Order after critical system code is secured."}'
                    f'</div></div>',
                    unsafe_allow_html=True)

                if not code_short.empty:
                    code_short_display = code_short[
                        ["Material_Code","Material_Name","UOM","Demand_Qty",
                         "Allocated_Qty","Shortfall_Qty","Fulfillment_Pct"]
                    ].copy()
                    plotly_mat_table(code_short_display,
                                     f"exec_other_{sel_tag}_{code}",
                                     height=65+len(code_short_display)*30)

            # Summary box
            all_short_df = tag_alloc[tag_alloc["Shortfall_Qty"]>0]
            total_to_order = all_short_df["Shortfall_Qty"].sum()

            st.markdown(f"""
            <div class="grand-box" style="margin-top:1rem;">
              <div style="font-family:'JetBrains Mono',monospace;font-size:.6rem;
                          letter-spacing:.14em;text-transform:uppercase;
                          color:#F59E0B;margin-bottom:.6rem;">
                Execution Summary — {sel_tag}</div>
              <div style="font-size:.82rem;color:var(--t2);line-height:1.8;">
                Critical system code <strong style="color:#F59E0B;">
                Code {sel_code} ({crit_sname})</strong> is at
                <strong style="color:{'#10B981' if crit_pct>=100 else '#EF4444'};">
                {crit_pct:.1f}%</strong> coverage.
                {"All critical materials are secured — proceed to other system codes." if crit_pct>=100
                  else f"Order {len(crit_short_df)} critical material(s) totalling "
                       f"<strong style='color:#EF4444;'>{crit_short:,.2f} units</strong> first."}
                {"" if other_short==0 else
                  f" Additionally, other system codes require "
                  f"<strong style='color:#F59E0B;'>{other_short:,.2f} units</strong>"
                  f" across {len(all_short_df[~all_short_df['Lining_System_Code'].isin([sel_code])])} material(s)."}
                <br>
                <strong style="color:var(--t0);">
                Total to order for full completion: {total_to_order:,.2f} units
                across {len(all_short_df)} material(s).</strong>
              </div>
            </div>""", unsafe_allow_html=True)

            if not all_short_df.empty:
                st.markdown("<br>", unsafe_allow_html=True)
                st.download_button(
                    f"⬇ Download Execution Order List — {sel_tag}",
                    data=excel_bytes(all_short_df[
                        ["Lining_System_Code","Lining_System_Short_Name",
                         "Material_Code","Material_Name","UOM",
                         "Demand_Qty","Allocated_Qty","Shortfall_Qty","Fulfillment_Pct"]
                    ].sort_values(["Lining_System_Code","Shortfall_Qty"],
                                  ascending=[True,False])),
                    file_name=f"execution_plan_{sel_tag.replace('/','-')}.xlsx",
                    mime="application/vnd.ms-excel",
                    use_container_width=True,
                )

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: DAILY CONSUMPTION ENTRY
# ═══════════════════════════════════════════════════════════════════════════════
with tab_consume:
    st.markdown('<div class="sec-hdr">📝 Daily Consumption Entry — Record actual site production</div>',
                unsafe_allow_html=True)

    if not db_available():
        st.error("Database not found. Run `python setup_db.py` first to initialise the database.")
        st.stop()

    # ── Cascading dropdowns: Location → Type → Equipment → System Code ────────
    st.markdown('<div class="sec-hdr">Step 1 — Select Work Location & Equipment</div>',
                unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        ce_loc = st.selectbox("📍 Location", options=[""] + LOCATION_ORDER,
                              key="ce_loc", label_visibility="visible")
    with col2:
        if ce_loc:
            type_opts = sorted(
                eq_master[eq_master["Location"]==ce_loc]["Type"].dropna().unique().tolist())
        else:
            type_opts = sorted(eq_master["Type"].dropna().unique().tolist())
        ce_type = st.selectbox("🏷 Type", options=[""] + type_opts,
                               key="ce_type", label_visibility="visible")
    with col3:
        eq_filter = eq_master.copy()
        if ce_loc:  eq_filter = eq_filter[eq_filter["Location"]==ce_loc]
        if ce_type: eq_filter = eq_filter[eq_filter["Type"]==ce_type]
        tag_opts = sorted(eq_filter["Equipment_Tag_No."].tolist())
        ce_tag = st.selectbox("🔩 Equipment Tag", options=[""] + tag_opts,
                              format_func=lambda t: "" if t=="" else _eq_label(t),
                              key="ce_tag", label_visibility="visible")
    with col4:
        if ce_tag:
            code_opts = sorted(
                equip_sc[equip_sc["Equipment_Tag_No."]==ce_tag]["Lining_System_Code"].unique(),
                key=lambda x: int(x))
            code_labels = [
                f"Code {c}  –  {equip_sc[(equip_sc['Equipment_Tag_No.']==ce_tag)&(equip_sc['Lining_System_Code']==c)]['Lining_System_Short_Name'].iloc[0]}"
                for c in code_opts]
        else:
            code_opts, code_labels = [], []
        ce_code_raw = st.selectbox("⚙️ System Code", options=[""] + code_labels,
                                   key="ce_code", label_visibility="visible")
        ce_code = ce_code_raw.split("  –  ")[0].replace("Code ","").strip() if ce_code_raw else ""

    # ── SQM Entry ──────────────────────────────────────────────────────────────
    if ce_tag and ce_code:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="sec-hdr">Step 2 — Enter SQM Completed Today</div>',
                    unsafe_allow_html=True)

        # Get remaining SQM for this (tag, code)
        sqm_row = sqm_ref[
            (sqm_ref["Equipment_Tag_No."]==ce_tag) &
            (sqm_ref["Lining_System_Code"]==ce_code)]
        total_sqm_orig  = float(sqm_row["Total_SQM_Original"].iloc[0]) if not sqm_row.empty else 0
        done_sqm_prev   = float(sqm_row["done_sqm"].iloc[0]) if not sqm_row.empty else 0
        remaining_sqm   = max(0.0, total_sqm_orig - done_sqm_prev)
        sname           = (equip_sc[(equip_sc["Equipment_Tag_No."]==ce_tag) &
                                    (equip_sc["Lining_System_Code"]==ce_code)]
                           ["Lining_System_Short_Name"].iloc[0]
                           if not sqm_row.empty else ce_code)

        sc1,sc2,sc3,sc4 = st.columns(4)
        sc1.metric("System Code",       sname)
        sc2.metric("Original SQM",      f"{total_sqm_orig:,.2f}")
        sc3.metric("Already Done SQM",  f"{done_sqm_prev:,.2f}")
        sc4.metric("Remaining SQM",     f"{remaining_sqm:,.2f}",
                   help="Remaining = Original − already completed")

        # ── Pre-load recipe before form ───────────────────────────────────────
        sc_recipe = recipe[recipe["Lining_System_Code"]==ce_code].copy()
        sc_recipe = sc_recipe.merge(
            inv[["Material_Code","Available_Qty","Ordered_Qty"]],
            on="Material_Code", how="left")
        sc_recipe["Available_Qty"] = sc_recipe["Available_Qty"].fillna(0)
        sc_recipe["Ordered_Qty"]   = sc_recipe["Ordered_Qty"].fillna(0)

        if sc_recipe.empty:
            st.warning(f"No recipe found for System Code {ce_code}.")
        else:
            st.markdown("<hr>", unsafe_allow_html=True)
            # ── All inputs in one form → no per-keystroke reruns ──────────────
            with st.form(key="ce_form", clear_on_submit=True):
                st.markdown('<div class="sec-hdr">Step 2 — Enter SQM Completed Today</div>',
                            unsafe_allow_html=True)
                col_date, col_sqm = st.columns(2)
                with col_date:
                    ce_date_form = st.date_input("📅 Work Date", value=date.today(),
                                                 key="form_ce_date")
                with col_sqm:
                    ce_sqm_form = st.number_input(
                        "SQM Completed Today",
                        min_value=0.0, max_value=float(remaining_sqm),
                        value=0.0, step=0.5, format="%.2f",
                        key="form_ce_sqm",
                        help=f"Maximum {remaining_sqm:,.2f} m² remaining")

                st.markdown('<div class="sec-hdr" style="margin-top:.8rem;">'
                            'Step 3 — Material Quantities Consumed</div>',
                            unsafe_allow_html=True)
                st.caption(
                    "Actual Consumed defaults to For_1_SQM × SQM if left at 0. "
                    "Override with actual site usage if different.")

                # Table header
                h1,h2,h3,h4,h5,h6,h7 = st.columns([2,3,1,1.5,1.5,1.5,1.5])
                for hdr, col in zip(
                    ["Code","Material Name","UOM","Available",
                     "For 1 SQM","Actual Consumed","On Order"],
                    [h1,h2,h3,h4,h5,h6,h7]
                ):
                    col.markdown(f"**{hdr}**")
                st.markdown("---")

                mat_inputs = {}
                for _, mrow in sc_recipe.iterrows():
                    mc    = str(mrow["Material_Code"])
                    for_1 = float(mrow.get("For_1_SQM", 0) or 0)
                    avail = float(mrow.get("Available_Qty", 0) or 0)
                    onord = float(mrow.get("Ordered_Qty",  0) or 0)
                    c1,c2,c3,c4,c5,c6,c7 = st.columns([2,3,1,1.5,1.5,1.5,1.5])
                    c1.markdown(f"<code>{mc}</code>", unsafe_allow_html=True)
                    c2.write(str(mrow.get("Material_Name", "")))
                    c3.write(str(mrow.get("UOM", "")))
                    c4.write(f"{avail:,.2f}")
                    c5.write(f"{for_1:,.4f}")
                    actual = c6.number_input(
                        "qty", min_value=0.0, value=0.0,
                        step=0.001, format="%.4f",
                        key=f"form_mat_{mc}",
                        label_visibility="collapsed")
                    c7.write(f"{onord:,.2f}")
                    mat_inputs[mc] = {
                        "material_name": str(mrow.get("Material_Name", "")),
                        "uom":           str(mrow.get("UOM", "")),
                        "for_1_sqm":     for_1,
                        "actual_input":  float(actual),
                    }

                ce_notes_form = st.text_area(
                    "📝 Notes (optional)",
                    placeholder="Weather conditions, issues, remarks…",
                    key="form_notes", height=70)

                submit_btn = st.form_submit_button(
                    "✅  Submit Daily Consumption",
                    use_container_width=False)

            # ── DB write — runs once on submit, not on every keypress ─────────
            if submit_btn:
                sqm_val  = st.session_state.get("form_ce_sqm",  0.0)
                date_val = st.session_state.get("form_ce_date", date.today())
                notes_val= st.session_state.get("form_notes",   "")
                if sqm_val <= 0:
                    st.error("Enter SQM Completed > 0 before submitting.")
                elif sqm_val > remaining_sqm + 0.001:
                    st.error(f"SQM ({sqm_val:.2f}) exceeds remaining ({remaining_sqm:.2f}).")
                else:
                    try:
                        conn = get_db()
                        cur  = conn.cursor()
                        n_updated = 0
                        for mc, vals in mat_inputs.items():
                            # Auto-calc expected; use actual override if > 0
                            expected_qty  = round(vals["for_1_sqm"] * sqm_val, 4)
                            consumed_qty  = (vals["actual_input"]
                                             if vals["actual_input"] > 0
                                             else expected_qty)
                            cur.execute("""
                                INSERT INTO consumption_log
                                  (entry_date, equipment_tag, lining_system_code,
                                   lining_system_name, sqm_completed, material_code,
                                   material_name, uom, expected_qty, consumed_qty, notes)
                                VALUES (?,?,?,?,?,?,?,?,?,?,?)
                            """, (str(date_val), ce_tag, ce_code, sname,
                                  sqm_val, mc, vals["material_name"], vals["uom"],
                                  expected_qty, consumed_qty, notes_val))
                            if consumed_qty > 0:
                                cur.execute("""
                                    UPDATE inventory
                                    SET available_qty = MAX(0, available_qty - ?)
                                    WHERE material_code = ?
                                """, (consumed_qty, mc))
                                n_updated += 1
                        # Update SQM progress — deduct done_sqm from remaining
                        cur.execute("""
                            UPDATE sqm_progress
                            SET done_sqm = done_sqm + ?
                            WHERE equipment_tag = ? AND lining_system_code = ?
                        """, (sqm_val, ce_tag, ce_code))
                        conn.commit()
                        conn.close()
                        # Clear cache so Dashboard / Location / Total Overview refresh
                        st.cache_data.clear()
                        st.success(
                            f"✅ {sqm_val:.2f} SQM recorded for {ce_tag} · "
                            f"Code {ce_code} ({sname}).  "
                            f"{n_updated} material(s) deducted from inventory.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Database error: {e}")

    # ── Receipt Log ──────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    with st.expander("📦 Record Material Receipt (New Stock Received)", expanded=False):
        st.caption("Use this to log when new materials arrive on site. "
                   "Received quantities will be added to Available Inventory.")
        if db_available():
            with st.form(key="receipt_form", clear_on_submit=True):
                rc1, rc2, rc3, rc4 = st.columns(4)
                with rc1:
                    rc_date = st.date_input("📅 Receipt Date", value=date.today(),
                                            key="rc_date")
                with rc2:
                    mat_opts_r = inv["Material_Code"].tolist()
                    rc_mat = st.selectbox(
                        "🧪 Material",
                        options=mat_opts_r,
                        format_func=lambda m: f"{m}  –  "
                            f"{inv.set_index('Material_Code')['Material_Name'].get(m,m)[:30]}",
                        key="rc_mat")
                with rc3:
                    rc_qty = st.number_input("Qty Received", min_value=0.0,
                                             value=0.0, step=1.0, format="%.2f",
                                             key="rc_qty")
                with rc4:
                    rc_notes = st.text_input("Notes / PO Ref.", key="rc_notes",
                                             placeholder="PO number, supplier…")
                rc_submit = st.form_submit_button("➕  Record Receipt",
                                                  use_container_width=False)

            if rc_submit:
                if rc_qty <= 0:
                    st.error("Please enter a quantity > 0.")
                else:
                    try:
                        conn = get_db()
                        cur  = conn.cursor()
                        # Insert receipt log
                        cur.execute("""
                            INSERT INTO receipt_log
                              (entry_date, material_code, material_name,
                               uom, received_qty, notes)
                            SELECT ?, ?, material_name, uom, ?, ?
                            FROM inventory WHERE material_code = ?
                        """, (str(rc_date), rc_mat, rc_qty, rc_notes, rc_mat))
                        # Add to Available_Qty
                        cur.execute("""
                            UPDATE inventory
                            SET available_qty = available_qty + ?
                            WHERE material_code = ?
                        """, (rc_qty, rc_mat))
                        conn.commit()
                        conn.close()
                        st.cache_data.clear()
                        mat_name = inv.set_index("Material_Code")["Material_Name"].get(
                            rc_mat, rc_mat)
                        st.success(f"✅ Received {rc_qty:,.2f} units of {mat_name} "
                                   f"added to inventory.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")

            # Show recent receipts as collapsible per-material items
            if db_available():
                conn = get_db()
                rlog = pd.read_sql(
                    "SELECT entry_date AS Date, material_code AS Code, "
                    "material_name AS Material, uom AS UOM, "
                    "received_qty AS \"Received Qty\", notes AS Notes, "
                    "submitted_at AS \"Received At\" "
                    "FROM receipt_log ORDER BY submitted_at DESC LIMIT 100", conn)
                conn.close()
                if not rlog.empty:
                    st.markdown('<div class="sec-hdr" style="margin-top:.6rem;">'
                                'Recent Receipts</div>', unsafe_allow_html=True)
                    # Group by material code for cleaner display
                    for mat_code, grp in rlog.groupby("Code", sort=False):
                        mat_name = grp["Material"].iloc[0]
                        total_recv = grp["Received Qty"].sum()
                        latest = grp["Date"].iloc[0]
                        with st.expander(
                            f"📦  {mat_code}  ·  {mat_name}  ·  "
                            f"Total received: {total_recv:,.2f}  ·  Last: {latest}",
                            expanded=False,
                        ):
                            show_cols = ["Date","UOM","Received Qty","Notes","Received At"]
                            st.dataframe(
                                grp[show_cols].reset_index(drop=True),
                                use_container_width=True, hide_index=True)
                    st.download_button(
                        "⬇ Download Receipt Log",
                        data=excel_bytes(rlog,
                                         report_title="Material Receipt Log — Smart Material Estimator",
                                         add_grand_total=True),
                        file_name=f"receipt_log_{date.today()}.xlsx",
                        mime="application/vnd.ms-excel",
                        key="dl_receipt_log")

    # ── Consumption History ────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    with st.expander("📜 View Consumption History", expanded=False):
        conn = get_db()
        log_df = pd.read_sql("""
            SELECT entry_date AS Date,
                   equipment_tag AS Equipment,
                   lining_system_code AS Code,
                   lining_system_name AS System,
                   sqm_completed AS "SQM Done",
                   material_code AS Material,
                   material_name AS "Material Name",
                   uom AS UOM,
                   expected_qty AS "Expected Qty",
                   consumed_qty AS "Consumed Qty",
                   notes AS Notes,
                   submitted_at AS "Submitted At"
            FROM consumption_log
            ORDER BY submitted_at DESC
            LIMIT 200
        """, conn)
        conn.close()

        if log_df.empty:
            st.info("No consumption entries yet.")
        else:
            # Filter controls
            fh1, fh2 = st.columns(2)
            with fh1:
                log_tags = ["All"] + sorted(log_df["Equipment"].unique().tolist())
                fh_tag   = st.selectbox("Filter by Equipment", log_tags, key="log_tag")
            with fh2:
                log_dates = ["All"] + sorted(log_df["Date"].unique().tolist(), reverse=True)
                fh_date   = st.selectbox("Filter by Date",     log_dates, key="log_date")

            df_show = log_df.copy()
            if fh_tag  != "All": df_show = df_show[df_show["Equipment"]==fh_tag]
            if fh_date != "All": df_show = df_show[df_show["Date"]==fh_date]

            st.dataframe(df_show, use_container_width=True, hide_index=True,
                         height=min(600, 50+len(df_show)*35))
            st.download_button(
                "⬇ Download Consumption Log",
                data=excel_bytes(df_show,
                                 report_title="Daily Consumption Log — Smart Material Estimator",
                                 add_grand_total=True),
                file_name=f"consumption_log_{date.today()}.xlsx",
                mime="application/vnd.ms-excel",
                use_container_width=False)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5 · TOTAL OVERVIEW  (Master Filterable Datatable)
# ═══════════════════════════════════════════════════════════════════════════════
with tab5:
    st.markdown('<div class="sec-hdr">📈 Total Overview — Master Equipment & Material Table</div>',
                unsafe_allow_html=True)

    # ── Build master table: one row per (Equipment, System Code) ─────────────
    # Join eq_master + sqm_ref + demand aggregation
    sqm_ref_ov = equip_sc[["Equipment_Tag_No.","Lining_System_Code",
                            "Lining_System_Short_Name","Total_SQM_Original",
                            "done_sqm","Total_SQM"]].drop_duplicates()

    # Total demand per (tag, code) from recipe × SQM
    dm_agg = dm.groupby(["Equipment_Tag_No.","Lining_System_Code"],
                         as_index=False).agg(
        Total_Demand_Qty=("Demand_Qty","sum"))
    dm_agg = dm_agg.merge(
        inv[["Material_Code","Available_Qty"]].groupby("Material_Code",as_index=False).first(),
        how="cross")  # we need per-code shortfall

    # Simpler: shortfall per (tag,code) = demand − min(demand, available)
    # Use cascade_allocate with all tags in file order for true shortfall
    _all_tags_ov = eq_master["Equipment_Tag_No."].tolist()
    _alloc_ov    = cascade_allocate(_all_tags_ov)

    sc_shortfall = _alloc_ov.groupby(
        ["Equipment_Tag_No.","Lining_System_Code"], as_index=False
    ).agg(
        Shortfall_Qty=("Shortfall_Qty","sum"),
        Demand_Qty   =("Demand_Qty","sum"),
        Allocated_Qty=("Allocated_Qty","sum"),
    )

    # Master table
    master = sqm_ref_ov.rename(columns={
        "Total_SQM_Original":"Total_SQM",
        "done_sqm":          "Done_SQM",
        "Total_SQM":         "Remaining_SQM",
    })
    # Merge equipment master (with all columns from Data Input sheet)
    master = master.merge(
        eq_master[["Equipment_Tag_No.","Name","Description","Location","Type",
                   "Lining_Systems","Lining_Type","Material_Spec","Design"]],
        on="Equipment_Tag_No.", how="left")

    # Merge equipment_sc for Lining_Area (surface area per row, not summed)
    lining_area_ref = equip_sc[
        ["Equipment_Tag_No.","Lining_System_Code","Total_SQM_Original"]
    ].drop_duplicates().rename(columns={"Total_SQM_Original":"Lining_Area_SQM"})
    master = master.merge(lining_area_ref,
                          on=["Equipment_Tag_No.","Lining_System_Code"], how="left")

    master = master.merge(sc_shortfall[["Equipment_Tag_No.","Lining_System_Code",
                                         "Shortfall_Qty","Demand_Qty","Allocated_Qty"]],
                          on=["Equipment_Tag_No.","Lining_System_Code"], how="left")
    master["Shortfall_Qty"] = master["Shortfall_Qty"].fillna(0)
    master["Fulfillment_%"] = (master["Allocated_Qty"] /
        master["Demand_Qty"].replace(0,np.nan) * 100).fillna(100).clip(0,100).round(1)

    # Add serial number
    master = master.reset_index(drop=True)
    master.insert(0, "S.No", master.index + 1)

    display_master = master[[
        "S.No","Equipment_Tag_No.","Name","Description","Type","Location",
        "Lining_Systems","Lining_System_Code","Lining_System_Short_Name",
        "Lining_Type","Material_Spec","Design",
        "Total_SQM","Lining_Area_SQM","Done_SQM","Remaining_SQM",
        "Demand_Qty","Allocated_Qty","Shortfall_Qty","Fulfillment_%"
    ]].copy()
    display_master.columns = [
        "S.No","Equipment No","Name","Description","Type","Location",
        "Lining System+","System Code","System Name",
        "Lining Type","Material Spec.","Design",
        "Total SQM","Lining Area SQM","Done SQM","Remaining SQM",
        "Total Demand","Allocated","Shortfall Qty","Fulfil %"
    ]

    # ── Filter controls ───────────────────────────────────────────────────────
    st.markdown('<div class="sec-hdr">🎛 Filters</div>', unsafe_allow_html=True)
    ff1,ff2,ff3,ff4 = st.columns(4)
    with ff1:
        f_loc_ov = st.multiselect("📍 Location",
            options=LOCATION_ORDER, default=LOCATION_ORDER, key="ov_loc")
    with ff2:
        type_opts_ov = sorted(display_master["Type"].dropna().unique().tolist())
        f_type_ov = st.multiselect("🏷 Type",
            options=type_opts_ov, default=type_opts_ov, key="ov_type")
    with ff3:
        codes_ov = sorted(display_master["System Code"].unique().tolist(), key=int)
        f_code_ov = st.multiselect("⚙️ System Code",
            options=codes_ov, default=codes_ov, key="ov_code")
    with ff4:
        status_opts = ["All","Fully Ready (100%)","Partial (50-99%)","Blocked (<50%)"]
        f_status_ov = st.selectbox("📊 Status", options=status_opts, key="ov_status")

    # Apply filters
    filtered_master = display_master[
        display_master["Location"].isin(f_loc_ov) &
        display_master["Type"].isin(f_type_ov) &
        display_master["System Code"].isin(f_code_ov)
    ].copy()
    if f_status_ov == "Fully Ready (100%)":
        filtered_master = filtered_master[filtered_master["Fulfil %"] >= 100]
    elif f_status_ov == "Partial (50-99%)":
        filtered_master = filtered_master[(filtered_master["Fulfil %"] >= 50) &
                                          (filtered_master["Fulfil %"] < 100)]
    elif f_status_ov == "Blocked (<50%)":
        filtered_master = filtered_master[filtered_master["Fulfil %"] < 50]

    # Renumber after filter
    filtered_master = filtered_master.reset_index(drop=True)
    filtered_master["S.No"] = filtered_master.index + 1

    # ── Dynamic summary KPIs ─────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    ov1,ov2,ov3,ov4,ov5,ov6 = st.columns(6)
    ov1.metric("Rows (filtered)",    len(filtered_master),
               help="Number of (Equipment, System Code) pairs in current filter.")
    ov2.metric("Total SQM",          f'{filtered_master["Total SQM"].sum():,.1f}',
               help="Sum of original SQM for filtered rows.")
    ov3.metric("Done SQM",           f'{filtered_master["Done SQM"].sum():,.1f}',
               help="SQM already completed (from daily consumption entries).")
    ov4.metric("Remaining SQM",      f'{filtered_master["Remaining SQM"].sum():,.1f}',
               help="SQM still to be completed = Total − Done.")
    ov5.metric("Total Shortfall Qty",f'{filtered_master["Shortfall Qty"].sum():,.0f}',
               help="Total material units short across all filtered rows.")
    ov6.metric("Avg Coverage",
               f'{filtered_master["Fulfil %"].mean():.1f}%' if len(filtered_master) else "0%",
               help="Average fulfillment % across filtered (Equipment, System Code) pairs.")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Colour-coded master table ─────────────────────────────────────────────
    def _style_master(row):
        pct = row["Fulfil %"]
        if pct >= 100:  bg,tc = "rgba(16,185,129,.1)","#10B981"
        elif pct >= 90: bg,tc = "rgba(249,115,22,.1)","#F97316"
        elif pct >= 80: bg,tc = "rgba(234,179,8,.1)", "#EAB308"
        else:           bg,tc = "rgba(239,68,68,.1)", "#EF4444"
        styles = [f"background-color:{bg}"] * len(row)
        ci = list(row.index).index("Fulfil %")
        styles[ci] = f"background-color:{bg};color:{tc};font-weight:700"
        return styles

    styled_master = (filtered_master.style
        .apply(_style_master, axis=1)
        .format({
            "Total SQM":     "{:,.2f}",
            "Done SQM":      "{:,.2f}",
            "Remaining SQM": "{:,.2f}",
            "Total Demand":  "{:,.2f}",
            "Allocated":     "{:,.2f}",
            "Shortfall Qty": "{:,.2f}",
            "Fulfil %":      "{:.1f}%",
        }))
    st.dataframe(styled_master, use_container_width=True, hide_index=True,
                 height=min(700, 50 + len(filtered_master)*35),
                 key="ov_master_tbl")

    # ── Per-System-Code material detail (expandable) ─────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="sec-hdr">🔬 Material Detail by System Code</div>',
                unsafe_allow_html=True)

    for code in sorted(filtered_master["System Code"].unique().tolist(), key=int):
        sc_dm = dm[dm["Lining_System_Code"]==code]
        if sc_dm.empty: continue
        sname   = filtered_master[filtered_master["System Code"]==code]["System Name"].iloc[0]
        sc_sqm  = filtered_master[filtered_master["System Code"]==code]["Total SQM"].sum()
        done_sq = filtered_master[filtered_master["System Code"]==code]["Done SQM"].sum()
        sc_mat  = sc_dm.groupby(["Material_Code","Material_Name","UOM"],
                                 as_index=False)["Demand_Qty"].sum()
        sc_mat  = sc_mat.merge(inv[["Material_Code","Available_Qty"]],
                               on="Material_Code", how="left")
        sc_mat["Available_Qty"] = sc_mat["Available_Qty"].fillna(0)
        sc_mat["Shortfall"]     = (sc_mat["Demand_Qty"]-sc_mat["Available_Qty"]).clip(lower=0).round(3)
        sc_mat["Coverage_%"]    = (
            sc_mat["Available_Qty"].clip(upper=sc_mat["Demand_Qty"])
            / sc_mat["Demand_Qty"].replace(0,np.nan)*100
        ).fillna(100).clip(0,100).round(1)
        # SQM achievable based on MINIMUM material coverage (bottleneck)
        sc_cov_min = sc_mat["Coverage_%"].min() if len(sc_mat) else 100
        sc_cov_avg = (sc_mat["Available_Qty"].clip(upper=sc_mat["Demand_Qty"]).sum() /
                      sc_mat["Demand_Qty"].sum()*100) if sc_mat["Demand_Qty"].sum()>0 else 100
        sc_can      = sc_sqm * min(1.0, sc_cov_avg/100)
        dot = "🟢" if sc_cov_avg>=100 else "🟠" if sc_cov_avg>=90 else "🟡" if sc_cov_avg>=80 else "🔴"

        with st.expander(
            f"{dot}  Code {code}  ·  {sname}  ·  "
            f"{sc_can:,.1f}/{sc_sqm:,.1f} SQM  ·  Done: {done_sq:,.1f}  ·  {sc_cov_avg:.1f}%",
            expanded=False,
        ):
            m1,m2,m3,m4,m5 = st.columns(5)
            m1.metric("System Code", f"Code {code}")
            m2.metric("Short Name",  sname)
            m3.metric("Total SQM",   f"{sc_sqm:,.2f}")
            m4.metric("Done SQM",    f"{done_sq:,.2f}",
                      help="SQM completed via Daily Consumption entries.")
            m5.metric("SQM Can Do",  f"{sc_can:,.2f}  ({sc_cov_avg:.1f}%)",
                      help="Achievable SQM with current inventory balance.")

            mat_show = sc_mat[["Material_Code","Material_Name","UOM",
                                "Available_Qty","Demand_Qty","Shortfall",
                                "Coverage_%"]].copy()
            mat_show.columns = ["Code","Material Name","UOM",
                                 "Available","Total Demand","Shortfall","Coverage %"]
            def _style_ov_det(row):
                pct=row["Coverage %"]
                if pct>=100:  bg,tc="rgba(16,185,129,.1)","#10B981"
                elif pct>=90: bg,tc="rgba(249,115,22,.1)","#F97316"
                elif pct>=80: bg,tc="rgba(234,179,8,.1)", "#EAB308"
                else:         bg,tc="rgba(239,68,68,.1)", "#EF4444"
                styles=[f"background-color:{bg}"]*len(row)
                ci=list(row.index).index("Coverage %")
                styles[ci]=f"background-color:{bg};color:{tc};font-weight:700"
                return styles
            st.dataframe(
                mat_show.style.apply(_style_ov_det,axis=1).format({
                    "Available":"{:,.3f}","Total Demand":"{:,.3f}",
                    "Shortfall":"{:,.3f}","Coverage %":"{:.1f}%"}),
                use_container_width=True, hide_index=True,
                height=65+len(mat_show)*35, key=f"ov_det_{code}")

    # ── Downloads ─────────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            "⬇ Download Filtered Master Table",
            data=excel_bytes(filtered_master,
                             report_title=f"Total Overview — {date.today()}",
                             add_grand_total=True),
            file_name=f"total_overview_{date.today()}.xlsx",
            mime="application/vnd.ms-excel",
            use_container_width=True)
    with dl2:
        if db_available():
            conn = get_db()
            full_log = pd.read_sql("SELECT * FROM consumption_log ORDER BY submitted_at DESC", conn)
            conn.close()
            st.download_button(
                "⬇ Download Full Consumption Log",
                data=excel_bytes(full_log,
                                 report_title="Full Consumption Log — Smart Material Estimator"),
                file_name=f"consumption_log_full_{date.today()}.xlsx",
                mime="application/vnd.ms-excel",
                use_container_width=True)